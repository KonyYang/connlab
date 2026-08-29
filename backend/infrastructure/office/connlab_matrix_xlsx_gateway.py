"""Read ConnLab-generated Matrix XLSX workbooks without trusting hidden state blindly."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from backend.modules.test_plan import (
    MatrixGroupPreview,
    MatrixParseResult,
    MatrixRowPreview,
    MatrixStepPreview,
    parse_step_tokens,
)
from backend.modules.test_plan.connlab_matrix_xlsx_format import (
    METADATA_SCHEMA,
    METADATA_SHEET_NAME,
    METADATA_VERSION,
    canonical_fingerprint,
    visible_matrix_fingerprint,
    visible_matrix_payload,
)


_VISIBLE_PREFIX = ("Test Item", "Section", "Test Method", "Condition", "Requirement")
_DAY_FALLBACK_WARNING = (
    "Day is not present in the visible ConnLab Matrix workbook; imported rows default "
    "to 0 days. Review is optional."
)
_STALE_METADATA_WARNING = (
    "Saved workbook details did not match the visible Matrix table; visible values "
    "were imported and Day defaulted to 0."
)


class ConnLabMatrixXlsxGateway:
    """Parse the strict ConnLab visible layout and optional trusted metadata."""

    def read(self, source_path: Path) -> MatrixParseResult:
        try:
            workbook = load_workbook(
                source_path,
                data_only=False,
                read_only=False,
                keep_links=False,
            )
        except Exception:
            return MatrixParseResult(
                groups=(), blockers=("Cannot read the selected ConnLab Matrix workbook.",)
            )
        try:
            sheet = self._visible_sheet(workbook)
            if sheet is None:
                return MatrixParseResult(
                    groups=(),
                    blockers=("No worksheet matches the ConnLab Matrix export structure.",),
                )
            visible, problem = self._read_visible(sheet)
            if problem:
                return MatrixParseResult(groups=(), blockers=(problem,))
            metadata, metadata_status = self._trusted_metadata(workbook, visible)
            return self._build_result(visible, metadata, metadata_status)
        finally:
            workbook.close()

    @staticmethod
    def _visible_sheet(workbook):
        for sheet in workbook.worksheets:
            values = tuple(_text(sheet.cell(1, column).value) for column in range(1, 6))
            if values == _VISIBLE_PREFIX:
                return sheet
        return None

    @staticmethod
    def _read_visible(sheet) -> tuple[dict[str, Any], str | None]:
        max_column = sheet.max_column
        if max_column > 70 or sheet.max_row > 516:
            return {}, "ConnLab Matrix workbook exceeds the supported 64 Groups or 512 test rows."
        if max_column < 7 or _text(sheet.cell(1, max_column).value) != "Notes":
            return {}, "ConnLab Matrix headers are incomplete or reordered."
        group_labels = [_text(sheet.cell(1, column).value) for column in range(6, max_column)]
        if any(not label for label in group_labels) or len(set(group_labels)) != len(group_labels):
            return {}, "ConnLab Matrix Group headers must be non-empty and unique."
        footer_rows: dict[str, int] = {}
        for row_number in range(2, sheet.max_row + 1):
            label = _text(sheet.cell(row_number, 1).value)
            if label in {"Sample size", "Time", "Fee"}:
                footer_rows[label] = row_number
        sample_row = footer_rows.get("Sample size")
        time_row = footer_rows.get("Time")
        fee_row = footer_rows.get("Fee")
        if not sample_row or time_row != sample_row + 1 or fee_row != sample_row + 2:
            return {}, "ConnLab Matrix footer must contain consecutive Sample size, Time, and Fee rows."
        if any(
            _text(sheet.cell(row_number, column).value)
            for row_number in range(fee_row + 1, sheet.max_row + 1)
            for column in range(1, max_column + 1)
        ):
            return {}, "ConnLab Matrix workbook contains unexpected content after the Fee row."
        for row in sheet.iter_rows(min_row=1, max_row=fee_row, min_col=1, max_col=max_column):
            if any(cell.data_type == "f" for cell in row):
                return {}, "ConnLab Matrix import does not accept formulas in the visible table."
        rows: list[dict[str, Any]] = []
        for row_number in range(2, sample_row):
            values = [_text(sheet.cell(row_number, column).value) for column in range(1, max_column)]
            if not any(values):
                continue
            rows.append(
                {
                    "source_row_index": row_number,
                    "test_item": values[0],
                    "section": values[1],
                    "test_method": values[2],
                    "condition": values[3],
                    "requirement": values[4],
                    "steps": values[5:],
                    "note": _text(sheet.cell(row_number, max_column).value),
                }
            )
        if not rows:
            return {}, "ConnLab Matrix workbook does not contain any test rows."
        sample_sizes = [_text(sheet.cell(sample_row, column).value) for column in range(6, max_column)]
        time_displays = [_text(sheet.cell(time_row, column).value) for column in range(6, max_column)]
        fees = [_text(sheet.cell(fee_row, column).value) for column in range(6, max_column)]
        fingerprint_payload = visible_matrix_payload(
            group_labels=group_labels,
            rows=rows,
            sample_sizes=sample_sizes,
            time_displays=time_displays,
            fees=fees,
        )
        return {
            "group_labels": group_labels,
            "rows": rows,
            "sample_sizes": sample_sizes,
            "time_displays": time_displays,
            "fees": fees,
            "fingerprint": visible_matrix_fingerprint(fingerprint_payload),
        }, None

    @staticmethod
    def _trusted_metadata(workbook, visible: dict[str, Any]) -> tuple[dict[str, Any] | None, str]:
        if METADATA_SHEET_NAME not in workbook.sheetnames:
            return None, "missing"
        sheet = workbook[METADATA_SHEET_NAME]
        try:
            if (
                sheet.sheet_state not in {"hidden", "veryHidden"}
                or
                _text(sheet["A1"].value) != METADATA_SCHEMA
                or _text(sheet["B1"].value) != METADATA_VERSION
                or _text(sheet["A2"].value) != "visible_sha256"
                or _text(sheet["A3"].value) != "payload_json"
                or _text(sheet["A4"].value) != "payload_sha256"
                or _text(sheet["B2"].value) != visible["fingerprint"]
            ):
                return None, "stale"
            raw_payload = sheet["B3"].value
            if not isinstance(raw_payload, str) or len(raw_payload) > 2_000_000:
                return None, "stale"
            payload = json.loads(raw_payload)
            if not isinstance(payload, dict):
                return None, "stale"
            if _text(sheet["B4"].value) != canonical_fingerprint(payload):
                return None, "stale"
            groups = payload.get("groups")
            rows = payload.get("rows")
            if not isinstance(groups, list) or not isinstance(rows, list):
                return None, "stale"
            if len(groups) != len(visible["group_labels"]) or len(rows) != len(visible["rows"]):
                return None, "stale"
            if not _metadata_values_are_supported(payload):
                return None, "stale"
            return payload, "trusted"
        except (TypeError, ValueError):
            return None, "stale"

    @staticmethod
    def _build_result(
        visible: dict[str, Any], metadata: dict[str, Any] | None, metadata_status: str
    ) -> MatrixParseResult:
        metadata_groups = metadata.get("groups", []) if metadata else []
        metadata_rows = metadata.get("rows", []) if metadata else []
        group_keys = [
            _text(item.get("group_key")) if isinstance(item, dict) else ""
            for item in metadata_groups
        ]
        if metadata is None or any(not key for key in group_keys) or len(set(group_keys)) != len(group_keys):
            group_keys = [f"group_{index}" for index in range(1, len(visible["group_labels"]) + 1)]
            metadata = None
            metadata_status = "stale" if metadata_status != "missing" else "missing"
        rows: list[MatrixRowPreview] = []
        steps_by_group: dict[str, list[MatrixStepPreview]] = {key: [] for key in group_keys}
        token_warnings: list[str] = []
        for index, raw_row in enumerate(visible["rows"]):
            tokens = {
                group_key: raw_row["steps"][group_index]
                for group_index, group_key in enumerate(group_keys)
                if raw_row["steps"][group_index]
            }
            day = "0"
            if metadata is not None:
                item = metadata_rows[index]
                if isinstance(item, dict):
                    day = _text(item.get("day_expression")) or "0"
            row = MatrixRowPreview(
                source_row_index=raw_row["source_row_index"],
                test_item=raw_row["test_item"],
                source_section=raw_row["section"] or None,
                group_tokens=tokens,
                method=raw_row["test_method"] or None,
                condition=raw_row["condition"] or None,
                requirement=raw_row["requirement"] or None,
                detail_extraction_status="xlsx_visible_table",
                day_expression=day,
            )
            rows.append(row)
            for group_key, token_text in tokens.items():
                parsed_tokens, warnings = parse_step_tokens(token_text)
                token_warnings.extend(
                    f"{visible['group_labels'][group_keys.index(group_key)]}: {warning}"
                    for warning in warnings
                )
                for token in parsed_tokens:
                    steps_by_group[group_key].append(
                        MatrixStepPreview(
                            sequence=token.sequence,
                            raw_token=token.raw_token,
                            suffix_note=token.suffix_note,
                            test_item=row.test_item,
                            source_section=row.source_section,
                            source_table_index=1,
                            source_row_index=row.source_row_index,
                            condition_summary=row.condition,
                            method_summary=row.method,
                            judgement_criteria=row.requirement,
                        )
                    )
        groups: list[MatrixGroupPreview] = []
        for index, (group_key, label) in enumerate(zip(group_keys, visible["group_labels"])):
            expression = visible["sample_sizes"][index] or None
            metadata_group = metadata_groups[index] if metadata is not None else None
            groups.append(
                MatrixGroupPreview(
                    group_key=group_key,
                    group_label=label,
                    source_table_index=1,
                    extraction_status="xlsx_metadata" if metadata is not None else "xlsx_visible_table",
                    steps=tuple(steps_by_group[group_key]),
                    sample_size=_positive_int(expression),
                    sample_quantity_expression=expression,
                    sample_note=(
                        _text(metadata_group.get("sample_note")) or None
                        if isinstance(metadata_group, dict)
                        else None
                    ),
                )
            )
        warnings: list[str] = []
        if metadata is None:
            warnings.append(_DAY_FALLBACK_WARNING if metadata_status == "missing" else _STALE_METADATA_WARNING)
        warnings.extend(token_warnings)
        schedule = metadata.get("schedule") if metadata is not None else None
        if not isinstance(schedule, dict):
            schedule = None
        else:
            schedule = {str(key): (_text(value) or None) for key, value in schedule.items()}
        return MatrixParseResult(
            groups=tuple(groups),
            warnings=tuple(warnings),
            blockers=(),
            selected_table_index=1,
            rows=tuple(rows),
            schedule=schedule,
        )


def _text(value: object | None) -> str:
    return str(value or "").strip()


def _positive_int(value: str | None) -> int | None:
    try:
        parsed = int(str(value).strip())
    except (TypeError, ValueError):
        return None
    return parsed if parsed > 0 else None


def _metadata_values_are_supported(payload: dict[str, Any]) -> bool:
    groups = payload.get("groups")
    rows = payload.get("rows")
    schedule = payload.get("schedule", {})
    if not isinstance(groups, list) or not isinstance(rows, list) or not isinstance(schedule, dict):
        return False
    for group in groups:
        if not isinstance(group, dict):
            return False
        if not 1 <= len(_text(group.get("group_key"))) <= 128:
            return False
        if len(_text(group.get("sample_note"))) > 2048:
            return False
    for row in rows:
        if not isinstance(row, dict) or len(_text(row.get("day_expression"))) > 64:
            return False
    allowed_schedule = {
        "post_test_buffer_days",
        "sample_received_date",
        "planned_test_start_date",
        "planned_test_complete_date",
        "estimated_completion_date",
    }
    return set(schedule).issubset(allowed_schedule) and all(
        len(_text(value)) <= 64 for value in schedule.values()
    )
