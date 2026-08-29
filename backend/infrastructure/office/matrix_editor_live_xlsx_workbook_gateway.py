"""Render a Matrix Editor live snapshot to an in-memory XLSX workbook."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from backend.application.matrix_editor_live_xlsx_export_service import (
    MatrixEditorLiveXlsxExportProjection,
)
from backend.modules.test_plan.connlab_matrix_xlsx_format import (
    METADATA_SCHEMA,
    METADATA_SHEET_NAME,
    METADATA_VERSION,
    canonical_fingerprint,
    canonical_json,
    visible_matrix_fingerprint,
    visible_matrix_payload,
)


class MatrixEditorLiveXlsxWorkbookGateway:
    """Build the reference-shaped workbook without external templates."""

    def render(self, projection: MatrixEditorLiveXlsxExportProjection) -> bytes:
        """Return one macro-free workbook as bytes."""
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet"
        headers = [
            "Test Item", "Section", "Test Method", "Condition", "Requirement",
            *(group.group_label for group in projection.groups), "Notes",
        ]
        sheet.append(headers)
        for column in range(6, 6 + len(projection.groups)):
            self._literalize(sheet.cell(1, column))
        for row in projection.rows:
            sheet.append([
                row.test_item, row.section, row.test_method, row.condition,
                row.requirement, *(cell.step_text for cell in row.cells), None,
            ])
            for column in range(1, 6 + len(projection.groups)):
                self._literalize(sheet.cell(sheet.max_row, column))
        for label, values in (
            ("Sample size", [group.sample_size or None for group in projection.groups]),
            ("Time", [group.time_display or None for group in projection.groups]),
            ("Fee", [None for _ in projection.groups]),
        ):
            sheet.append([label, None, None, None, None, *values, None])
            if label != "Fee":
                for column in range(6, 6 + len(projection.groups)):
                    self._literalize(sheet.cell(sheet.max_row, column))
        self._format(sheet)
        self._write_metadata(workbook, projection)
        stream = BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    @staticmethod
    def _write_metadata(workbook, projection: MatrixEditorLiveXlsxExportProjection) -> None:
        visible = visible_matrix_payload(
            group_labels=(group.group_label for group in projection.groups),
            rows=(
                {
                    "test_item": row.test_item,
                    "section": row.section,
                    "test_method": row.test_method,
                    "condition": row.condition,
                    "requirement": row.requirement,
                    "steps": [cell.step_text for cell in row.cells],
                    "note": "",
                }
                for row in projection.rows
            ),
            sample_sizes=(group.sample_size for group in projection.groups),
            time_displays=(group.time_display for group in projection.groups),
            fees=("" for _ in projection.groups),
        )
        payload = {
            "groups": [
                {
                    "group_id": group.group_id,
                    "group_key": group.group_key,
                    "sample_note": group.sample_note,
                }
                for group in projection.groups
            ],
            "rows": [
                {
                    "row_id": row.row_id,
                    "day_expression": row.day_expression,
                }
                for row in projection.rows
            ],
            "schedule": (
                {
                    key: getattr(projection.schedule, key)
                    for key in (
                        "post_test_buffer_days",
                        "sample_received_date",
                        "planned_test_start_date",
                        "planned_test_complete_date",
                        "estimated_completion_date",
                    )
                }
                if projection.schedule is not None
                else {}
            ),
        }
        metadata = workbook.create_sheet(METADATA_SHEET_NAME)
        metadata.sheet_state = "veryHidden"
        metadata["A1"] = METADATA_SCHEMA
        metadata["B1"] = METADATA_VERSION
        metadata["A2"] = "visible_sha256"
        metadata["B2"] = visible_matrix_fingerprint(visible)
        payload_json = canonical_json(payload)
        metadata["A3"] = "payload_json"
        metadata["B3"] = payload_json
        metadata["A4"] = "payload_sha256"
        metadata["B4"] = canonical_fingerprint(payload)
        for cell in (
            metadata["A1"], metadata["B1"], metadata["A2"], metadata["B2"],
            metadata["A3"], metadata["B3"], metadata["A4"], metadata["B4"],
        ):
            cell.data_type = "s"

    @staticmethod
    def _literalize(cell) -> None:
        """Keep user-editable text from being interpreted as an Excel formula."""
        if isinstance(cell.value, str):
            cell.data_type = "s"

    @staticmethod
    def _format(sheet) -> None:
        gray = PatternFill("solid", fgColor="CCCCCC")
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(name="Calibri", size=11)
                cell.alignment = alignment
                cell.border = border
        for cell in sheet[1]:
            cell.fill = gray
        for row_number in range(2, sheet.max_row + 1):
            sheet.cell(row_number, 1).fill = gray
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 8
        sheet.column_dimensions["D"].width = 20
        sheet.column_dimensions["E"].width = 20
