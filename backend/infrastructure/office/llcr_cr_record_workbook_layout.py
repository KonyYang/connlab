"""Shared fixed sheet layout for confirmed and draft LLCR/CR workbooks."""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

LLCR_CR_RECORD_LAYOUT_V1 = {
    "summary_sheet": "Record Summary",
    "llcr_sheet": "LLCR Record",
    "cr_sheet": "CR Record",
    "record_headers": (
        "Type", "Group", "Source Step", "Sample", "Contact ID", "Contact Label",
        "Initial", "After", "Final", "Result", "Remarks",
    ),
}

LLCR_CR_RECORD_LAYOUT_V2 = {
    "summary_sheet": "Record Summary",
    "summary_headers": (
        "Type", "Point category", "Group", "Samples", "Points / sample",
        "Stages", "Generated rows",
    ),
}

_BORDER = Border(
    left=Side(style="thin", color="B8C2CE"),
    right=Side(style="thin", color="B8C2CE"),
    top=Side(style="thin", color="B8C2CE"),
    bottom=Side(style="thin", color="B8C2CE"),
)

_MACRO_BORDER = Border(
    left=Side(style="thin", color="808080"),
    right=Side(style="thin", color="808080"),
    top=Side(style="thin", color="808080"),
    bottom=Side(style="thin", color="808080"),
)
_MACRO_HEADER_FILL = PatternFill("solid", fgColor="DCDCDC")
_MACRO_INPUT_FILL = PatternFill("solid", fgColor="FFFFC8")
_MACRO_STATS_FILL = PatternFill("solid", fgColor="99CCFF")
_MACRO_UNUSED_FILL = PatternFill("solid", fgColor="E7E6E6")


def write_record_sheet(sheet, sections, *, banner: str | None = None) -> None:
    """Write fixed Group-Step blocks, manual cells, formulas, and widths."""
    row_index = 1
    if banner:
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
        sheet.cell(1, 1, banner)
        sheet.cell(1, 1).font = Font(bold=True, size=14)
        sheet.cell(1, 1).fill = PatternFill("solid", fgColor="E8EEF6")
        row_index = 3
    for section in sections:
        sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=11)
        sheet.cell(row_index, 1, f"{_display_type(section.record_type)} | {section.group_label} | Step {section.source_step}")
        sheet.cell(row_index, 1).font = Font(bold=True)
        sheet.cell(row_index, 1).fill = PatternFill("solid", fgColor="D8E0EA")
        sheet.cell(row_index + 1, 1, "Samples")
        sheet.cell(row_index + 1, 2, section.sample_count)
        sheet.cell(row_index + 1, 3, "Readings / sample")
        sheet.cell(row_index + 1, 4, section.readings_per_sample)
        header_row = row_index + 2
        write_header_row(sheet, header_row, LLCR_CR_RECORD_LAYOUT_V1["record_headers"])
        first_row = header_row + 1
        for offset, record in enumerate(section.rows):
            current = first_row + offset
            values = (_display_type(section.record_type), section.group_label, section.source_step, record.sample_index, record.contact_id, record.contact_label)
            for column, value in enumerate(values, start=1):
                sheet.cell(current, column, value)
        last_row = first_row + len(section.rows) - 1
        summary_row = last_row + 1
        sheet.cell(summary_row, 6, "Statistics")
        sheet.cell(summary_row, 7, _average_formula("G", first_row, last_row))
        sheet.cell(summary_row, 8, _average_formula("H", first_row, last_row))
        sheet.cell(summary_row, 9, _average_formula("I", first_row, last_row))
        sheet.cell(summary_row, 10, _result_formula(first_row, last_row))
        row_index = summary_row + 2
    set_column_widths(sheet, (14, 22, 16, 10, 16, 24, 14, 14, 14, 16, 28))


def write_header_row(sheet, row: int, headers: tuple[str, ...]) -> None:
    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(row, column, value)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8EEF6")
        cell.alignment = Alignment(horizontal="center")


def set_column_widths(sheet, widths: tuple[int, ...]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width


def _average_formula(column: str, first_row: int, last_row: int) -> str:
    return f'=IF(COUNT({column}{first_row}:{column}{last_row})=0,"",AVERAGE({column}{first_row}:{column}{last_row}))'


def _result_formula(first_row: int, last_row: int) -> str:
    return f'=IF(COUNTA(J{first_row}:J{last_row})=0,"",COUNTIF(J{first_row}:J{last_row},"PASS")&"/"&COUNTA(J{first_row}:J{last_row}))'


def _display_type(record_type: str) -> str:
    return "CR" if record_type == "cr_specified_current" else "LLCR"


def write_specialized_category_sheet(
    sheet,
    sections,
    *,
    record_type: str,
    delta_r_enabled: bool,
) -> None:
    """Write one point-category sheet with one shared correction input set."""
    sections = tuple(sections)
    if not sections:
        return
    first = sections[0]
    points = []
    seen = set()
    for row in first.rows:
        if row.contact_id not in seen:
            seen.add(row.contact_id)
            points.append(row.contact_id)

    max_columns = max(
        10,
        max(
            3 + sum(
                _stage_width(record_type, delta_r_enabled, stage_index)
                for stage_index, _stage in enumerate(section.stages)
            )
            for section in sections
        ),
    )
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_columns)
    sheet.cell(1, 1, f"{record_type.upper()} Test Record · {first.category_label or first.record_prefix or 'Points'}")
    sheet.cell(1, 1).font = Font(bold=True, size=16, color="17324D")
    sheet.cell(3, 1, "Point category")
    sheet.cell(3, 2, first.category_label or "")
    sheet.cell(4, 1, "Point prefix")
    sheet.cell(4, 2, first.record_prefix or "")
    sheet.cell(5, 1, "Point range")
    sheet.cell(5, 2, first.point_expression or "")
    sheet.cell(6, 1, "Correction")
    sheet.cell(6, 2, "Bulk Resistance" if record_type == "llcr" else "Bulk Voltage")
    write_header_row(
        sheet,
        8,
        ("Point ID", "Bulk Resistance (mΩ)" if record_type == "llcr" else "Bulk Voltage (mV)"),
    )
    for offset, point_id in enumerate(points, start=9):
        sheet.cell(offset, 1, point_id)
        sheet.cell(offset, 2, None)
        _apply_row_border(sheet, offset, 1, 2)
    bulk_first = 9
    bulk_last = 8 + len(points)
    next_row = bulk_last + 3
    for section in sections:
        next_row = _write_group_block(
            sheet,
            section,
            start_row=next_row,
            bulk_first=bulk_first,
            bulk_last=bulk_last,
            record_type=record_type,
            delta_r_enabled=delta_r_enabled,
        )
    sheet.freeze_panes = "D9"
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 22
    for column in range(4, max_columns + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 18


def write_macro_style_contact_resistance_category_sheet(
    sheet,
    sections,
    *,
    record_type: str,
    delta_r_enabled: bool,
    ltr_number: str | None,
) -> dict[tuple[str, int], tuple[int, int]]:
    """Write the LLCR/CR record shape used by the approved VBA workbook."""
    sections = tuple(sections)
    if not sections:
        return {}
    if record_type not in {"llcr", "cr"}:
        raise ValueError("Macro-style record type must be llcr or cr.")
    delta_r_enabled = record_type == "llcr" and delta_r_enabled
    max_samples = max(section.sample_count for section in sections)
    raw_start = 4
    raw_end = raw_start + max_samples - 1
    calculated_group = raw_end + 3
    calculated_step = calculated_group + 1
    calculated_point = calculated_group + 2
    corrected_start = calculated_group + 3
    delta_start = corrected_start + max_samples if delta_r_enabled else None
    stats_start = (
        delta_start + max_samples
        if delta_start is not None
        else corrected_start + max_samples
    )
    environment_start = stats_start + 4
    last_column = environment_start + 2

    _write_macro_bulk_table(
        sheet,
        record_type=record_type,
        test_current=_first_test_current(sections),
    )
    _write_macro_test_information(
        sheet,
        ltr_number=ltr_number,
        test_condition=_first_test_condition(sections),
    )
    _write_macro_record_header(
        sheet,
        max_samples=max_samples,
        calculated_group=calculated_group,
        corrected_start=corrected_start,
        delta_start=delta_start,
        stats_start=stats_start,
        environment_start=environment_start,
    )

    stats_cells: dict[tuple[str, int], tuple[int, int]] = {}
    current_row = 10
    for section in sections:
        points = _section_points(section)
        group_start = current_row
        initial_corrected_rows: list[int] = []
        for stage_index, stage in enumerate(section.stages):
            stage_start = current_row
            stage_end = stage_start + len(points) - 1
            stage_label = _record_stage_label(stage.label, stage_index, record_type)
            sheet.cell(stage_start, 2, stage_label)
            sheet.cell(stage_start, calculated_step, stage_label)
            _merge_vertical(sheet, stage_start, stage_end, 2)
            _merge_vertical(sheet, stage_start, stage_end, calculated_step)
            for point_offset, point_id in enumerate(points):
                row = stage_start + point_offset
                sheet.cell(row, 3, point_id)
                sheet.cell(row, calculated_point, point_id)
                if stage_index == 0:
                    initial_corrected_rows.append(row)
                for sample_index in range(1, max_samples + 1):
                    raw_column = raw_start + sample_index - 1
                    corrected_column = corrected_start + sample_index - 1
                    if sample_index > section.sample_count:
                        sheet.cell(row, raw_column).fill = _MACRO_UNUSED_FILL
                        sheet.cell(row, corrected_column).fill = _MACRO_UNUSED_FILL
                        if delta_start is not None:
                            sheet.cell(row, delta_start + sample_index - 1).fill = _MACRO_UNUSED_FILL
                        continue
                    raw_address = f"{get_column_letter(raw_column)}{row}"
                    corrected_address = f"{get_column_letter(corrected_column)}{row}"
                    if record_type == "cr":
                        corrected_formula = (
                            f'=IF(OR({raw_address}="",$B$6=""),"",'
                            f'({raw_address}-$B$5)/$B$6)'
                        )
                    else:
                        corrected_formula = (
                            f'=IF(OR({raw_address}="",$B$5=""),"",'
                            f'{raw_address}-$B$5)'
                        )
                    sheet.cell(row, corrected_column, corrected_formula)
                    if delta_start is not None:
                        delta_column = delta_start + sample_index - 1
                        if stage_index == 0:
                            formula = f'=IF({corrected_address}="","",{corrected_address})'
                        else:
                            initial_row = initial_corrected_rows[point_offset]
                            initial_address = (
                                f"{get_column_letter(corrected_column)}{initial_row}"
                            )
                            formula = (
                                f'=IF(OR({corrected_address}="",{initial_address}=""),"",'
                                f"{corrected_address}-{initial_address})"
                            )
                        sheet.cell(row, delta_column, formula)

            source_start = delta_start if delta_start is not None else corrected_start
            source_end = source_start + section.sample_count - 1
            source_range = (
                f"{get_column_letter(source_start)}{stage_start}:"
                f"{get_column_letter(source_end)}{stage_end}"
            )
            formulas = (
                f'=IF(COUNT({source_range})=0,"",MIN({source_range}))',
                f'=IF(COUNT({source_range})=0,"",MAX({source_range}))',
                f'=IF(COUNT({source_range})=0,"",AVERAGE({source_range}))',
                f'=IF(COUNT({source_range})<2,"",STDEV({source_range}))',
            )
            for offset, formula in enumerate(formulas):
                column = stats_start + offset
                sheet.cell(stage_start, column, formula)
                _merge_vertical(sheet, stage_start, stage_end, column)
            for column in range(environment_start, environment_start + 3):
                _merge_vertical(sheet, stage_start, stage_end, column)
            stats_cells[(section.confirmed_group_id, stage_index)] = (
                stage_start,
                stats_start,
            )
            current_row = stage_end + 1

        group_end = current_row - 1
        sheet.cell(group_start, 1, _group_display_label(section.group_label))
        sheet.cell(group_start, calculated_group, _group_display_label(section.group_label))
        _merge_vertical(sheet, group_start, group_end, 1)
        _merge_vertical(sheet, group_start, group_end, calculated_group)

    _format_macro_contact_resistance_sheet(
        sheet,
        end_row=current_row - 1,
        raw_end=raw_end,
        calculated_group=calculated_group,
        stats_start=stats_start,
        environment_start=environment_start,
        last_column=last_column,
        number_format="0.000" if record_type == "cr" else "0.0",
    )
    return stats_cells


def write_macro_style_contact_resistance_summary(
    sheet,
    category_outputs,
    *,
    parameter_labels: tuple[str, ...],
    record_type: str,
    delta_r_enabled: bool,
) -> None:
    """Write the VBA-style cross-category LLCR/CR statistics summary."""
    category_outputs = tuple(category_outputs)
    if not category_outputs:
        return
    if record_type not in {"llcr", "cr"}:
        raise ValueError("Macro-style summary type must be llcr or cr.")
    delta_r_enabled = record_type == "llcr" and delta_r_enabled
    canonical_sections = category_outputs[0][1]
    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=2)
    sheet.cell(1, 1, "Test Step")
    labels = _summary_category_labels(category_outputs, parameter_labels)
    for category_index, label in enumerate(labels):
        start_column = 3 + category_index * 4
        sheet.merge_cells(
            start_row=1,
            start_column=start_column,
            end_row=1,
            end_column=start_column + 3,
        )
        sheet.cell(1, start_column, label)
        for offset, header in enumerate(("Min", "Max", "Avg", "Stdev")):
            sheet.cell(2, start_column + offset, header)

    row = 3
    group_ranges: list[tuple[int, int]] = []
    for section in canonical_sections:
        group_start = row
        for stage_index, stage in enumerate(section.stages):
            sheet.cell(row, 2, _summary_stage_label(
                stage.label,
                stage_index,
                len(section.stages),
                delta_r_enabled,
                record_type,
            ))
            for category_index, (
                sheet_name,
                _sections,
                stats_cells,
                _fallback,
            ) in enumerate(category_outputs):
                target = stats_cells.get((section.confirmed_group_id, stage_index))
                if target is None:
                    continue
                stats_row, stats_column = target
                escaped_name = sheet_name.replace("'", "''")
                for offset in range(4):
                    source = f"{get_column_letter(stats_column + offset)}{stats_row}"
                    reference = f"'{escaped_name}'!{source}"
                    sheet.cell(
                        row,
                        3 + category_index * 4 + offset,
                        f'=IF({reference}="","",{reference})',
                    )
            row += 1
        sheet.cell(group_start, 1, _group_display_label(section.group_label))
        _merge_vertical(sheet, group_start, row - 1, 1)
        group_ranges.append((group_start, row - 1))

    last_column = 2 + len(category_outputs) * 4
    used = sheet.cell(row - 1, last_column).coordinate
    table = sheet[f"A1:{used}"]
    for cells in table:
        for cell in cells:
            cell.font = Font(name="Arial", size=9, bold=cell.row <= 2 or cell.column <= 2)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _MACRO_BORDER
            if cell.row <= 2 or cell.column <= 2:
                cell.fill = _MACRO_HEADER_FILL
            if cell.row >= 3 and cell.column >= 3:
                cell.number_format = "0.000" if record_type == "cr" else "0.0"
    for group_number, (group_start, group_end) in enumerate(group_ranges, start=1):
        if group_number % 2 == 0:
            for cells in sheet.iter_rows(
                min_row=group_start,
                max_row=group_end,
                min_col=3,
                max_col=last_column,
            ):
                for cell in cells:
                    cell.fill = _MACRO_INPUT_FILL
    sheet.column_dimensions["A"].width = 13
    sheet.column_dimensions["B"].width = 34
    for column in range(3, last_column + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 12
    sheet.freeze_panes = "C3"
    sheet.sheet_view.showGridLines = False


def _write_macro_bulk_table(
    sheet,
    *,
    record_type: str,
    test_current: float | None,
) -> None:
    unit = "unit:mV" if record_type == "cr" else "unit:mΩ"
    measurement = "Voltage" if record_type == "cr" else "Resistance"
    values = [
        (unit, measurement),
        ("bulk1", 0.0),
        ("bulk2", 0.0),
        ("bulk3", 0.0),
        ("Avg", '=IF(COUNT(B2:B4)=0,"",AVERAGE(B2:B4))'),
    ]
    if record_type == "cr":
        values.append(("Current(Unit:A)", test_current))
    for row, (label, value) in enumerate(values, start=1):
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, value)
        for column in (1, 2):
            cell = sheet.cell(row, column)
            cell.border = _MACRO_BORDER
            cell.font = Font(name="Arial", size=9, bold=column == 1 or row == 1)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.cell(row, 1).fill = _MACRO_HEADER_FILL
    sheet.cell(1, 2).fill = _MACRO_HEADER_FILL
    bulk_number_format = "0.000" if record_type == "cr" else "0.0"
    for row in range(2, 6):
        sheet.cell(row, 2).number_format = bulk_number_format


def _write_macro_test_information(
    sheet,
    *,
    ltr_number: str | None,
    test_condition: str,
) -> None:
    labels = ("LTR", "Tested By", "Checked by/Date", "Test Equipment ID", "Test Condition")
    values = (ltr_number or "", "", "", "", test_condition)
    for row, (label, value) in enumerate(zip(labels, values, strict=True), start=1):
        sheet.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        sheet.merge_cells(start_row=row, start_column=6, end_row=row, end_column=9)
        sheet.cell(row, 4, label)
        sheet.cell(row, 6, value)
        for column in range(4, 10):
            cell = sheet.cell(row, column)
            cell.border = _MACRO_BORDER
            cell.fill = _MACRO_INPUT_FILL
            cell.font = Font(name="Arial", size=9, bold=True)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_macro_record_header(
    sheet,
    *,
    max_samples: int,
    calculated_group: int,
    corrected_start: int,
    delta_start: int | None,
    stats_start: int,
    environment_start: int,
) -> None:
    sheet.merge_cells(start_row=9, start_column=1, end_row=9, end_column=2)
    sheet.cell(9, 1, "=A1")
    sheet.cell(9, 3, "S/N")
    for sample in range(1, max_samples + 1):
        sheet.cell(9, 3 + sample, f"{sample}#")
    sheet.merge_cells(
        start_row=9,
        start_column=calculated_group,
        end_row=9,
        end_column=calculated_group + 1,
    )
    sheet.cell(9, calculated_group, "unit:mΩ")
    sheet.cell(9, calculated_group + 2, "S/N")
    for sample in range(1, max_samples + 1):
        sheet.cell(9, corrected_start + sample - 1, f"{sample}#")
        if delta_start is not None:
            sheet.cell(9, delta_start + sample - 1, f"{sample}#ΔR")
    for offset, value in enumerate(("Min", "Max", "Avg", "Stdev")):
        sheet.cell(9, stats_start + offset, value)
    for offset, value in enumerate(("Test Date", "Amb Temp(°C)", "Rel. Hum.:%")):
        sheet.cell(9, environment_start + offset, value)


def _format_macro_contact_resistance_sheet(
    sheet,
    *,
    end_row: int,
    raw_end: int,
    calculated_group: int,
    stats_start: int,
    environment_start: int,
    last_column: int,
    number_format: str,
) -> None:
    for start_column, block_end in ((1, raw_end), (calculated_group, last_column)):
        for row in range(9, end_row + 1):
            for column in range(start_column, block_end + 1):
                cell = sheet.cell(row, column)
                cell.border = _MACRO_BORDER
                cell.font = Font(
                    name="Arial",
                    size=9,
                    bold=row == 9 or column in {start_column, start_column + 1},
                )
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if row == 9 or column in {start_column, start_column + 1}:
                    cell.fill = _MACRO_HEADER_FILL
                if column >= stats_start and column < environment_start:
                    cell.fill = _MACRO_STATS_FILL
                if column >= environment_start:
                    cell.fill = _MACRO_INPUT_FILL
                if row >= 10 and (
                    4 <= column <= raw_end
                    or calculated_group + 3 <= column < environment_start
                ):
                    cell.number_format = number_format
    widths = {
        1: 13,
        2: 34,
        3: 11,
        calculated_group: 13,
        calculated_group + 1: 34,
        calculated_group + 2: 11,
    }
    for column in range(4, last_column + 1):
        widths.setdefault(column, 12)
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "D10"
    sheet.sheet_view.showGridLines = False


def _section_points(section) -> tuple[str, ...]:
    points: list[str] = []
    seen: set[str] = set()
    for row in section.rows:
        if row.contact_id in seen:
            continue
        seen.add(row.contact_id)
        points.append(row.contact_id)
    return tuple(points)


def _first_test_condition(sections) -> str:
    for section in sections:
        for stage in section.stages:
            if stage.condition.strip():
                return stage.condition.strip()
    return ""


def _first_test_current(sections) -> float | None:
    for section in sections:
        for stage in section.stages:
            current = _number(stage.test_current_ampere)
            if current is not None:
                return current
    return None


def _record_stage_label(label: str, stage_index: int, record_type: str) -> str:
    value = label.strip()
    if stage_index == 0 and value == "Initial":
        return f"Initial {record_type.upper()}"
    if value == "Final":
        return f"Final {record_type.upper()}"
    return value


def _summary_stage_label(
    label: str,
    stage_index: int,
    stage_count: int,
    delta_r_enabled: bool,
    record_type: str,
) -> str:
    record_label = _record_stage_label(label, stage_index, record_type)
    if not delta_r_enabled or stage_index == 0:
        return record_label
    if stage_index == stage_count - 1 and label.strip() == "Final":
        return "Final ∆R"
    return f"∆R {record_label}"


def _summary_category_labels(category_outputs, parameter_labels: tuple[str, ...]) -> tuple[str, ...]:
    if len(category_outputs) == 1:
        return ("Statistics",)
    if len(parameter_labels) == len(category_outputs):
        return parameter_labels
    return tuple(fallback for _name, _sections, _stats, fallback in category_outputs)


def _group_display_label(value: str) -> str:
    label = value.strip()
    return label if label.lower().startswith("group") else f"Group {label}"


def _merge_vertical(sheet, start_row: int, end_row: int, column: int) -> None:
    if end_row > start_row:
        sheet.merge_cells(
            start_row=start_row,
            start_column=column,
            end_row=end_row,
            end_column=column,
        )


def _write_group_block(
    sheet,
    section,
    *,
    start_row: int,
    bulk_first: int,
    bulk_last: int,
    record_type: str,
    delta_r_enabled: bool,
) -> int:
    last_column = max(
        10,
        3 + sum(
            _stage_width(record_type, delta_r_enabled, stage_index)
            for stage_index, _stage in enumerate(section.stages)
        ),
    )
    sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=last_column)
    sheet.cell(start_row, 1, f"{section.group_label} · {section.sample_count} samples")
    sheet.cell(start_row, 1).font = Font(bold=True, color="FFFFFF")
    sheet.cell(start_row, 1).fill = PatternFill("solid", fgColor="3B5B7A")
    environment_row = start_row + 1
    for column, label in ((1, "Date"), (3, "Temperature (°C)"), (5, "RH (%)"), (7, "Tester"), (9, "Equipment")):
        sheet.cell(environment_row, column, label).font = Font(bold=True)
        sheet.cell(environment_row, column + 1, None)

    heading_row = start_row + 2
    current_row = heading_row + 1 if record_type == "cr" else None
    subheader_row = heading_row + (2 if record_type == "cr" else 1)
    data_start = subheader_row + 1
    write_header_row(sheet, subheader_row, ("Sample", "Point ID", "Point category"))
    corrected_columns: list[int] = []
    delta_columns: list[int] = []
    initial_corrected_column: int | None = None
    column = 4
    for stage_index, stage in enumerate(section.stages):
        stage_width = _stage_width(record_type, delta_r_enabled, stage_index)
        end_column = column + stage_width - 1
        sheet.merge_cells(
            start_row=heading_row, start_column=column,
            end_row=heading_row, end_column=end_column,
        )
        heading = sheet.cell(heading_row, column, f"{stage.label} · Step {stage.source_step}")
        heading.font = Font(bold=True)
        heading.fill = PatternFill("solid", fgColor="D8E6F3")
        heading.alignment = Alignment(horizontal="center")
        if record_type == "cr":
            assert current_row is not None
            sheet.cell(current_row, column, "Test current (A)").font = Font(bold=True)
            sheet.cell(current_row, column + 1, _number(stage.test_current_ampere))
            headers = ("Raw voltage (mV)", "Corrected (mΩ)")
        else:
            headers = (
                "Raw resistance (mΩ)",
                "Corrected (mΩ)",
                "ΔR (mΩ)",
            ) if delta_r_enabled and stage_index > 0 else (
                "Raw resistance (mΩ)", "Corrected (mΩ)"
            )
        for offset, value in enumerate(headers):
            cell = sheet.cell(subheader_row, column + offset, value)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E8EEF6")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        corrected_column = column + 1
        corrected_columns.append(corrected_column)
        if initial_corrected_column is None:
            initial_corrected_column = corrected_column
        if record_type == "llcr" and delta_r_enabled and stage_index > 0:
            delta_columns.append(column + 2)
        column = end_column + 1

    bulk_range = f"$A${bulk_first}:$B${bulk_last}"
    for offset, record in enumerate(section.rows):
        row = data_start + offset
        sheet.cell(row, 1, record.sample_index)
        sheet.cell(row, 2, record.contact_id)
        sheet.cell(row, 3, record.contact_label)
        column = 4
        for stage_index, _stage in enumerate(section.stages):
            stage_width = _stage_width(record_type, delta_r_enabled, stage_index)
            raw = f"{get_column_letter(column)}{row}"
            corrected_column = column + 1
            corrected = f"{get_column_letter(corrected_column)}{row}"
            correction_present = (
                f'COUNTIFS($A${bulk_first}:$A${bulk_last},$B{row},'
                f'$B${bulk_first}:$B${bulk_last},"<>")>0'
            )
            if record_type == "llcr":
                sheet.cell(
                    row,
                    corrected_column,
                    f'=IF(OR({raw}="",NOT({correction_present})),"",{raw}-VLOOKUP($B{row},{bulk_range},2,FALSE))',
                )
                if delta_r_enabled and stage_index > 0 and initial_corrected_column is not None:
                    initial = f"{get_column_letter(initial_corrected_column)}{row}"
                    sheet.cell(
                        row,
                        column + 2,
                        f'=IF(OR({corrected}="",{initial}=""),"",{corrected}-{initial})',
                    )
            else:
                assert current_row is not None
                current = f"${get_column_letter(corrected_column)}${current_row}"
                sheet.cell(
                    row,
                    corrected_column,
                    f'=IF(OR({raw}="",{current}="",NOT({correction_present})),"",({raw}-VLOOKUP($B{row},{bulk_range},2,FALSE))/{current})',
                )
            column += stage_width
        _apply_row_border(sheet, row, 1, last_column)

    data_last = data_start + len(section.rows) - 1
    stats_start = data_last + 2
    for offset, label in enumerate(("Minimum", "Maximum", "Average", "Stdev")):
        sheet.cell(stats_start + offset, 3, label).font = Font(bold=True)
    for numeric_column in corrected_columns + delta_columns:
        letter = get_column_letter(numeric_column)
        source = f"{letter}{data_start}:{letter}{data_last}"
        sheet.cell(stats_start, numeric_column, f'=IF(COUNT({source})=0,"",MIN({source}))')
        sheet.cell(stats_start + 1, numeric_column, f'=IF(COUNT({source})=0,"",MAX({source}))')
        sheet.cell(stats_start + 2, numeric_column, f'=IF(COUNT({source})=0,"",AVERAGE({source}))')
        sheet.cell(stats_start + 3, numeric_column, f'=IF(COUNT({source})<2,"",STDEV({source}))')
    return stats_start + 6


def _stage_width(record_type: str, delta_r_enabled: bool, stage_index: int) -> int:
    return 3 if record_type == "llcr" and delta_r_enabled and stage_index > 0 else 2


def _apply_row_border(sheet, row: int, first_column: int, last_column: int) -> None:
    for column in range(first_column, last_column + 1):
        sheet.cell(row, column).border = _BORDER


def _number(value: str | None):
    if value is None:
        return None
    try:
        return float(value)
    except ValueError:
        return None
