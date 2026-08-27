"""Macro-free XLSX writer for the Matrix Test Status workbook."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.application.test_status_workbook_projection import TestStatusProjection


_HEADER_FILL = PatternFill("solid", fgColor="FFC8C8C8")
_STATUS_FILL = PatternFill("solid", fgColor="FFADD8E6")
_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


class TestStatusWorkbookGateway:
    """Write one fresh Test Status workbook from the shared projection."""

    def write(self, *, output_path: Path, projection: TestStatusProjection) -> Path:
        target = Path(output_path)
        if target.suffix.lower() != ".xlsx":
            raise ValueError("Test Status output must be .xlsx.")
        if not target.parent.is_dir():
            raise FileNotFoundError(f"Output directory does not exist: {target.parent}")

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Test Status"
        sheet.append(["TEST", *(group.group_label for group in projection.groups)])
        for row in projection.rows:
            sheet.append(
                [
                    row.test_item,
                    *(
                        row.group_values.get(group.group_key, "")
                        for group in projection.groups
                    ),
                ]
            )
        sheet.append(
            ["Sample size", *(group.sample_quantity_expression for group in projection.groups)]
        )
        completion_row = sheet.max_row + 1
        sheet.cell(completion_row, 1, "Estimated completion date in Clarizen")
        sheet.cell(completion_row, 2, "")
        last_column = len(projection.groups) + 1
        if last_column > 2:
            sheet.merge_cells(
                start_row=completion_row,
                start_column=2,
                end_row=completion_row,
                end_column=last_column,
            )
        status_row = completion_row + 1
        sheet.cell(status_row, 1, "Status")
        for column in range(2, last_column + 1):
            sheet.cell(status_row, column, "No Start")

        for row in sheet.iter_rows(
            min_row=1,
            max_row=status_row,
            min_col=1,
            max_col=last_column,
        ):
            for cell in row:
                cell.border = _BORDER
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = _HEADER_FILL
        sample_row = completion_row - 1
        for row in sheet.iter_rows(
            min_row=sample_row,
            max_row=status_row,
            min_col=1,
            max_col=last_column,
        ):
            for cell in row:
                cell.fill = _STATUS_FILL
        sheet.column_dimensions["A"].width = 30
        for column in range(2, last_column + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 15
        sheet.row_dimensions[status_row].height = 60
        sheet.freeze_panes = "B2"
        workbook.save(target)
        return target
