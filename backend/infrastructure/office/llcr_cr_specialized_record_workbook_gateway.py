"""Macro-free openpyxl workbook writer for LLCR/CR specialized records."""

from __future__ import annotations

import os
from pathlib import Path
from tempfile import mkstemp
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from backend.application.confirmed_matrix_llcr_cr_record_projection import (
    LlcrCrRecordProjection,
)
from backend.infrastructure.office.llcr_cr_record_workbook_layout import (
    write_macro_style_contact_resistance_category_sheet,
    write_macro_style_contact_resistance_summary,
)



class LlcrCrSpecializedRecordWorkbookGateway:
    """Write one fresh macro-free `.xlsx` workbook from a ready projection."""

    def write(self, *, output_path: Path, projection: LlcrCrRecordProjection) -> Path:
        """Create the fixed workbook layout without mutating source authority."""
        target = Path(output_path)
        if target.suffix.lower() != ".xlsx":
            raise ValueError("LLCR/CR specialized record output must be .xlsx.")
        if projection.status not in {"ready", "complete", "partial_compatible"}:
            raise ValueError("A ready LLCR/CR record preview is required for generation.")
        if projection.record_type not in {"llcr", "cr"}:
            raise ValueError("A single LLCR or CR record type is required for generation.")
        if not target.parent.is_dir():
            raise FileNotFoundError(f"Output directory does not exist: {target.parent}")

        workbook = Workbook()
        workbook.remove(workbook.active)
        self._write_macro_style_workbook(workbook, projection)
        warning_ranges = _number_stored_as_text_warning_ranges(workbook)
        workbook.save(target)
        _suppress_number_stored_as_text_warnings(target, warning_ranges)
        return target

    def _write_macro_style_workbook(
        self,
        workbook: Workbook,
        projection: LlcrCrRecordProjection,
    ) -> None:
        summary = workbook.create_sheet("Summary")
        grouped: dict[str, list] = {}
        for section in projection.sections:
            if section.record_type != projection.record_type:
                raise ValueError("Record projection mixes LLCR and CR sections.")
            key = section.category_id or section.record_prefix or section.category_label or "Points"
            grouped.setdefault(key, []).append(section)
        used_names = {"Summary"}
        category_outputs = []
        for sections in grouped.values():
            sheet_name = _sheet_name(
                sections[0].record_prefix or sections[0].category_label or "Points",
                used_names,
            )
            used_names.add(sheet_name)
            sheet = workbook.create_sheet(sheet_name)
            stats_cells = write_macro_style_contact_resistance_category_sheet(
                sheet,
                sections,
                record_type=projection.record_type,
                delta_r_enabled=projection.delta_r_enabled,
                ltr_number=projection.ltr_number,
            )
            category_outputs.append((
                sheet_name,
                tuple(sections),
                stats_cells,
                sections[0].category_label or sections[0].record_prefix or "Statistics",
            ))
        write_macro_style_contact_resistance_summary(
            summary,
            category_outputs,
            parameter_labels=projection.summary_parameter_labels,
            record_type=projection.record_type,
            delta_r_enabled=projection.delta_r_enabled,
        )


def _sheet_name(value: str, used: set[str]) -> str:
    import re

    base = re.sub(r"[\\/*?:\[\]]+", "_", value.strip())[:31] or "Points"
    candidate = base
    suffix = 2
    while candidate in used:
        tail = f"_{suffix}"
        candidate = f"{base[:31 - len(tail)]}{tail}"
        suffix += 1
    return candidate


def _number_stored_as_text_warning_ranges(workbook: Workbook) -> dict[str, str]:
    ranges: dict[str, str] = {}
    for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
        sn_columns = [
            column
            for column in range(1, sheet.max_column + 1)
            if sheet.cell(9, column).value == "S/N"
        ]
        if not sn_columns or sheet.max_row < 10:
            continue
        ranges[f"xl/worksheets/sheet{sheet_index}.xml"] = " ".join(
            f"{get_column_letter(column)}10:{get_column_letter(column)}{sheet.max_row}"
            for column in sn_columns
        )
    return ranges


def _suppress_number_stored_as_text_warnings(
    target: Path,
    warning_ranges: dict[str, str],
) -> None:
    if not warning_ranges:
        return
    descriptor, temporary_name = mkstemp(
        dir=target.parent,
        prefix=f".{target.stem}-",
        suffix=target.suffix,
    )
    os.close(descriptor)
    temporary = Path(temporary_name)
    try:
        with ZipFile(target, "r") as source, ZipFile(
            temporary,
            "w",
            compression=ZIP_DEFLATED,
        ) as destination:
            destination.comment = source.comment
            for member in source.infolist():
                payload = source.read(member.filename)
                sqref = warning_ranges.get(member.filename)
                if sqref is not None:
                    payload = _insert_number_stored_as_text_ignored_error(payload, sqref)
                destination.writestr(member, payload)
        os.replace(temporary, target)
    finally:
        temporary.unlink(missing_ok=True)


def _insert_number_stored_as_text_ignored_error(sheet_xml: bytes, sqref: str) -> bytes:
    ignored_errors = (
        f'<ignoredErrors><ignoredError sqref="{sqref}" '
        'numberStoredAsText="1"/></ignoredErrors>'
    ).encode("utf-8")
    tail_markers = (
        b"<smartTags",
        b"<drawing",
        b"<legacyDrawing",
        b"<legacyDrawingHF",
        b"<picture",
        b"<oleObjects",
        b"<controls",
        b"<webPublishItems",
        b"<tableParts",
        b"<extLst",
        b"</worksheet>",
    )
    positions = [position for marker in tail_markers if (position := sheet_xml.find(marker)) >= 0]
    if not positions:
        raise ValueError("Generated worksheet XML has no valid ignored-errors insertion point.")
    insert_at = min(positions)
    return sheet_xml[:insert_at] + ignored_errors + sheet_xml[insert_at:]
