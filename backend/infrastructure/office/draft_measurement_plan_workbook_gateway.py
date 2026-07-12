"""Macro-free layout writer for review-only measurement-plan draft workbooks."""

from __future__ import annotations

from pathlib import Path
from datetime import datetime, timezone

from openpyxl import Workbook

from backend.application.draft_measurement_plan_workbook_projection import (
    DraftMeasurementPlanWorkbookProjection,
)
from backend.infrastructure.office.llcr_cr_specialized_record_workbook_gateway import (
    LLCR_CR_RECORD_LAYOUT_V1,
)
from backend.infrastructure.office.llcr_cr_record_workbook_layout import (
    write_record_sheet,
)


class DraftMeasurementPlanWorkbookGateway:
    """Write only valid draft/review projections to a fresh `.xlsx` file."""

    def write(self, *, output_path: Path, projection: DraftMeasurementPlanWorkbookProjection) -> Path:
        if output_path.suffix.lower() != ".xlsx":
            raise ValueError("Draft contact measurement output must be .xlsx.")
        if not projection.generate_allowed or not projection.output_label:
            raise ValueError("A valid draft workbook preview is required for generation.")
        if not output_path.parent.is_dir():
            raise FileNotFoundError(f"Output directory does not exist: {output_path.parent}")
        workbook = Workbook()
        workbook.remove(workbook.active)
        summary = workbook.create_sheet(LLCR_CR_RECORD_LAYOUT_V1["summary_sheet"])
        llcr = workbook.create_sheet(LLCR_CR_RECORD_LAYOUT_V1["llcr_sheet"])
        cr = workbook.create_sheet(LLCR_CR_RECORD_LAYOUT_V1["cr_sheet"])
        _write_summary(summary, projection)
        _write_records(llcr, projection, "llcr")
        _write_records(cr, projection, "cr_specified_current")
        workbook.save(output_path)
        return output_path


def _write_summary(sheet, projection: DraftMeasurementPlanWorkbookProjection) -> None:
    sheet["A1"] = projection.output_label
    sheet["A2"] = "LLCR/CR Draft Measurement Plan Workbook"
    metadata = (
        ("Project ID", projection.project_id),
        ("Plan revision", projection.revision_id),
        ("Plan fingerprint", projection.revision_fingerprint),
        ("Source Matrix", projection.matrix_id),
        ("Matrix binding fingerprint", projection.matrix_binding_fingerprint),
        ("Matrix revision", projection.matrix_revision),
        ("Preview fingerprint", projection.preview_fingerprint),
        ("Layout version", "LLCR_CR_RECORD_LAYOUT_V1"),
        ("Review diagnostics", len(projection.diagnostics)),
        ("Generated UTC", datetime.now(timezone.utc).isoformat()),
        ("Generated rows", projection.row_count),
    )
    for row, (label, value) in enumerate(metadata, start=3):
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, value)
    sheet["A15"] = "DRAFT ONLY. Not a confirmed Matrix or formal record."
    for index, section in enumerate(projection.sections, start=18):
        sheet.cell(index, 1, section.record_type.upper())
        sheet.cell(index, 2, section.group_label)
        sheet.cell(index, 3, section.source_step)
        sheet.cell(index, 4, len(section.rows))


def _write_records(sheet, projection: DraftMeasurementPlanWorkbookProjection, kind: str) -> None:
    write_record_sheet(
        sheet,
        tuple(item for item in projection.sections if item.record_type == kind),
        banner=f"{projection.output_label} | Draft measurement plan",
    )
