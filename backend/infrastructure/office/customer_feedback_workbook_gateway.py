"""Customer Feedback workbook generation gateway."""

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook


class CustomerFeedbackWorkbookGatewayError(RuntimeError):
    """Raised when Customer Feedback workbook generation cannot proceed."""


class CustomerFeedbackWorkbookGateway:
    """Generate Customer Feedback workbooks through a safe Office boundary."""

    def generate(
        self,
        *,
        template_path: Path,
        output_path: Path,
        identity: dict[str, str],
    ) -> tuple[Path, tuple[str, ...]]:
        """Copy the Customer Feedback template and fill known header labels."""
        template = Path(template_path)
        target = Path(output_path)
        if template.suffix.lower() != ".xlsx":
            raise CustomerFeedbackWorkbookGatewayError(
                f"Customer Feedback template must be an .xlsx file: {template}"
            )
        if target.suffix.lower() != ".xlsx":
            raise CustomerFeedbackWorkbookGatewayError(
                f"Customer Feedback output must be an .xlsx file: {target}"
            )
        if not template.is_file():
            raise CustomerFeedbackWorkbookGatewayError(
                f"Customer Feedback template does not exist: {template}"
            )
        if template.resolve() == target.resolve():
            raise CustomerFeedbackWorkbookGatewayError(
                "Customer Feedback output must not overwrite the source template."
            )
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(template, target)
        warnings = _fill_identity_fields(target, identity)
        return target, tuple(warnings)


def _fill_identity_fields(path: Path, identity: dict[str, str]) -> list[str]:
    """Fill simple label/value header cells in the copied workbook."""
    rules = {
        "ltr_number": _FieldRule(
            aliases=("work request no", "work request number", "ltr number", "dl number", "project number", "project no"),
            offsets=(1,),
            special_offsets={"work request no": 3, "work request number": 3},
            required_label="Work Request No.",
        ),
        "product_name": _FieldRule(
            aliases=("project details", "product name", "product description", "product"),
            offsets=(1,),
            special_offsets={"project details": 2},
            required_label="Project Details",
        ),
        "requestor": _FieldRule(
            aliases=("customer name", "requestor", "requester", "requested by"),
            offsets=(1,),
            special_offsets={"customer name": 2},
        ),
        "phone": _FieldRule(
            aliases=("telephone no", "phone", "telephone", "tel"),
            offsets=(1,),
        ),
        "location": _FieldRule(
            aliases=("site", "mfg site", "manufacturing site"),
            offsets=(1,),
            special_offsets={"site": 3},
        ),
        "received_date": _FieldRule(
            aliases=("from date", "date lab received samples", "received date"),
            offsets=(1,),
            special_offsets={"from date": 2},
        ),
        "estimated_completion_date": _FieldRule(
            aliases=("to date", "estimated completion date", "completion date"),
            offsets=(1,),
        ),
        "lab": _FieldRule(
            aliases=("ges team", "lab performing the tests", "lab", "testing lab"),
            offsets=(1,),
            special_offsets={"ges team": 2},
        ),
        "email": _FieldRule(
            aliases=("e-mail of requestor", "email", "e mail", "requestor email"),
            offsets=(1,),
        ),
        "project_leader": _FieldRule(
            aliases=("project leader", "engineer", "owner"),
            offsets=(1,),
        ),
        "test_item": _FieldRule(
            aliases=("test item", "tests to be performed", "requested testing"),
            offsets=(1,),
        ),
    }
    workbook = load_workbook(path)
    warnings: list[str] = []
    changed = False
    try:
        for field_key, value in identity.items():
            if not value or field_key not in rules:
                continue
            rule = rules[field_key]
            location = _find_target_cell(workbook, rule)
            if location is None:
                message = f"Customer Feedback header field not found: {rule.required_label or field_key}"
                if rule.required_label:
                    raise CustomerFeedbackWorkbookGatewayError(message)
                warnings.append(message)
                continue
            sheet_name, row, column = location
            sheet = workbook[sheet_name]
            target = sheet.cell(row=row, column=column)
            if str(target.value or "").startswith("="):
                message = f"Customer Feedback target cell contains a formula: {field_key}"
                if rule.required_label:
                    raise CustomerFeedbackWorkbookGatewayError(message)
                warnings.append(message)
                continue
            if target.value != value:
                target.value = value
                changed = True
        if changed:
            workbook.save(path)
    finally:
        workbook.close()
    return warnings


class _FieldRule:
    """Customer Feedback header placement rule."""

    def __init__(
        self,
        *,
        aliases: tuple[str, ...],
        offsets: tuple[int, ...],
        special_offsets: dict[str, int] | None = None,
        required_label: str | None = None,
    ) -> None:
        self.aliases = aliases
        self.offsets = offsets
        self.special_offsets = special_offsets or {}
        self.required_label = required_label


def _find_target_cell(workbook, rule: _FieldRule) -> tuple[str, int, int] | None:
    """Find the target value cell for one Customer Feedback identity rule."""
    label_location = _find_label_cell(workbook, rule.aliases)
    if label_location is None:
        return None
    sheet_name, row, column, matched_alias = label_location
    sheet = workbook[sheet_name]
    offsets = (
        (rule.special_offsets[matched_alias],)
        if matched_alias in rule.special_offsets
        else rule.offsets
    )
    for offset in offsets:
        target_column = column + offset
        if target_column <= sheet.max_column + 3:
            return sheet_name, row, target_column
    return None


def _find_label_cell(workbook, aliases: tuple[str, ...]) -> tuple[str, int, int, str] | None:
    normalized_aliases = {_normalize(alias) for alias in aliases}
    for sheet in workbook.worksheets:
        max_row = min(sheet.max_row, 30)
        max_column = min(sheet.max_column, 12)
        for row in range(1, max_row + 1):
            for column in range(1, max_column + 1):
                value = sheet.cell(row=row, column=column).value
                normalized_value = _normalize(str(value or ""))
                for alias in normalized_aliases:
                    if normalized_value == alias or normalized_value.startswith(f"{alias} "):
                        return sheet.title, row, column, alias
    return None


def _normalize(value: str) -> str:
    return " ".join(
        "".join(ch.lower() if ch.isalnum() else " " for ch in value).split()
    )
