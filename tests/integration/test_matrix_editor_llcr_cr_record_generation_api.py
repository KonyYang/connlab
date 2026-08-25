from pathlib import Path
from types import SimpleNamespace

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.api.dependencies import (
    get_matrix_editor_llcr_cr_record_generation_service,
)
from backend.api.main import app
from backend.application.contact_point_profile_confirmed_consumer_adapter import (
    EffectiveConfirmedPointProfile,
)
from backend.application.matrix_editor_llcr_cr_record_generation_service import (
    MatrixEditorLlcrCrRecordGenerationService,
)
from backend.infrastructure.files.llcr_cr_specialized_record_artifact_store import (
    LlcrCrSpecializedRecordArtifactStore,
)
from backend.infrastructure.office.llcr_cr_specialized_record_workbook_gateway import (
    LlcrCrSpecializedRecordWorkbookGateway,
)
from backend.domain.enums import LtrStatus


def test_matrix_editor_llcr_download_uses_current_ui_draft_without_confirmed_matrix(
    tmp_path: Path,
) -> None:
    response = _post_current_ui_draft(tmp_path, "5", None)

    assert response.status_code == 200
    assert "Unconfirmed" in response.headers["content-disposition"]
    output = tmp_path / "matrix-editor-llcr-preview.xlsx"
    output.write_bytes(response.content)
    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == ["Summary", "SIG"]
    sheet = workbook["SIG"]
    assert sheet["F1"].value == "DL-2026-05-999"
    assert sheet["F5"].value == "20 mV, 100 mA"
    assert sheet["D9"].value == "1#"
    assert sheet["K9"].value == "unit:mΩ"


def test_matrix_editor_llcr_download_accepts_footnoted_sample_quantity(
    tmp_path: Path,
) -> None:
    response = _post_current_ui_draft(
        tmp_path,
        "3(a)",
        "(a) Male connector and Female connector",
    )

    assert response.status_code == 200
    output = tmp_path / "matrix-editor-llcr-footnoted-preview.xlsx"
    output.write_bytes(response.content)
    sheet = load_workbook(output, data_only=False)["SIG"]
    assert [sheet.cell(9, column).value for column in range(4, 7)] == [
        "1#",
        "2#",
        "3#",
    ]
    assert sheet["I9"].value == "unit:mΩ"


def _post_current_ui_draft(
    tmp_path: Path,
    sample_quantity_expression: str,
    sample_note: str | None,
):
    generation = MatrixEditorLlcrCrRecordGenerationService(
        point_profile_adapter=_PointProfileAdapter(),
        workbook_gateway=LlcrCrSpecializedRecordWorkbookGateway(),
        artifact_store=LlcrCrSpecializedRecordArtifactStore(tmp_path / "generated"),
        ltr_store=_LtrStore(),
    )
    app.dependency_overrides[
        get_matrix_editor_llcr_cr_record_generation_service
    ] = lambda: generation
    try:
        response = TestClient(app).post(
            "/api/projects/P1/matrix-editor/llcr-cr-record-draft/generate",
            json={
                "source": "matrix_editor_current_ui_state",
                "record_type": "llcr",
                "groups": [
                    {
                        "group_key": "group_6",
                        "group_label": "6",
                        "sample_quantity_expression": sample_quantity_expression,
                        "sample_note": sample_note,
                    }
                ],
                "rows": [
                    {
                        "test_item": "Contact Resistance (Low Level)",
                        "section": "6.1",
                        "method": "EIA-364-23D",
                        "condition": "20 mV, 100 mA",
                        "requirement": "Initial <= 0.25 mOhm",
                        "group_values": {"group_6": "2,6"},
                    }
                ],
            },
        )
    finally:
        app.dependency_overrides.clear()
    return response


class _LtrStore:
    def list_by_project(self, project_id: str):
        assert project_id == "P1"
        return [
            SimpleNamespace(
                ltr_number="DL-2026-05-999",
                status=LtrStatus.REGISTERED,
                registered_on="2026-05-20",
            )
        ]


class _PointProfileAdapter:
    def get_effective(self, project_id: str) -> EffectiveConfirmedPointProfile:
        assert project_id == "P1"
        return EffectiveConfirmedPointProfile(
            status="confirmed",
            readings_per_sample="2",
            revision_id="profile-1",
            revision_sequence=1,
            fingerprint="profile-fingerprint",
            lineage="Confirmed Project Point Profile",
            message=None,
            categories=(
                {
                    "category_id": "signal",
                    "category_ordinal": 0,
                    "label": "Signal",
                    "count_per_sample": 2,
                    "record_prefix": "SIG",
                    "included": True,
                    "point_expression": "1-2",
                },
            ),
        )
