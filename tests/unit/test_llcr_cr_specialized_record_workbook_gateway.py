from __future__ import annotations

from dataclasses import replace

from openpyxl import load_workbook

from backend.application.confirmed_matrix_llcr_cr_record_projection import (
    LlcrCrRecordProjection,
    LlcrCrRecordRow,
    LlcrCrRecordSection,
    LlcrCrRecordStage,
)
from backend.infrastructure.office.llcr_cr_specialized_record_workbook_gateway import (
    LlcrCrSpecializedRecordWorkbookGateway,
)


def test_gateway_writes_macro_style_llcr_category_and_summary_formulas(tmp_path) -> None:
    output_path = tmp_path / "llcr-record.xlsx"
    gateway = LlcrCrSpecializedRecordWorkbookGateway()

    written = gateway.write(output_path=output_path, projection=_projection())

    assert written == output_path
    assert output_path.suffix == ".xlsx"
    workbook = load_workbook(output_path, data_only=False, keep_vba=False)
    assert workbook.sheetnames == ["Summary", "SIG"]
    summary = workbook["Summary"]
    assert summary["A1"].value == "Test Step"
    assert summary["C1"].value == "Statistics"
    assert [summary.cell(2, column).value for column in range(3, 7)] == [
        "Min", "Max", "Avg", "Stdev",
    ]
    assert summary["A3"].value == "Group 1"
    assert summary["B3"].value == "Initial LLCR"
    assert summary["B4"].value == "Final ∆R"
    assert summary["C3"].value == '=IF(\'SIG\'!L10="","",\'SIG\'!L10)'
    assert summary["C4"].value == '=IF(\'SIG\'!L12="","",\'SIG\'!L12)'
    sheet = workbook["SIG"]
    assert sheet["A1"].value == "unit:mΩ"
    assert sheet["B1"].value == "Resistance"
    assert [sheet.cell(row, 1).value for row in range(2, 6)] == [
        "bulk1", "bulk2", "bulk3", "Avg",
    ]
    assert sheet["B5"].value == '=IF(COUNT(B2:B4)=0,"",AVERAGE(B2:B4))'
    assert [sheet.cell(row, 4).value for row in range(1, 6)] == [
        "LTR", "Tested By", "Checked by/Date", "Test Equipment ID", "Test Condition",
    ]
    assert sheet["F1"].value == "DL-2025-11-073"
    assert sheet["F5"].value == "100 mA max"
    assert [sheet.cell(9, column).value for column in range(1, 19)] == [
        "=A1", None, "S/N", "1#", None, None,
        "unit:mΩ", None, "S/N", "1#", "1#ΔR",
        "Min", "Max", "Avg", "Stdev", "Test Date", "Amb Temp(°C)", "Rel. Hum.:%",
    ]
    assert sheet["A10"].value == "Group 1"
    assert sheet["B10"].value == "Initial LLCR"
    assert sheet["C10"].value == "SIG1"
    assert sheet["C11"].value == "SIG2"
    assert sheet["B12"].value == "Final LLCR"
    assert sheet["J10"].value == '=IF(OR(D10="",$B$5=""),"",D10-$B$5)'
    assert sheet["K10"].value == '=IF(J10="","",J10)'
    assert sheet["K12"].value == '=IF(OR(J12="",J10=""),"",J12-J10)'
    assert sheet["L10"].value == '=IF(COUNT(K10:K11)=0,"",MIN(K10:K11))'
    assert sheet["O10"].value == '=IF(COUNT(K10:K11)<2,"",STDEV(K10:K11))'
    assert sheet["D10"].number_format == "0.0"
    assert sheet["J10"].number_format == "0.0"
    assert sheet["P10"].number_format == "General"


def test_gateway_accepts_partial_compatible_llcr_projection(tmp_path) -> None:
    output_path = tmp_path / "partial-record.xlsx"
    projection = replace(
        _projection(),
        status="partial_compatible",
        measurement_plan_revision_id="revision-3",
        measurement_plan_revision_sequence=3,
        effective_measurement_plan_status="partial_compatible",
    )

    LlcrCrSpecializedRecordWorkbookGateway().write(
        output_path=output_path,
        projection=projection,
    )

    workbook = load_workbook(output_path, data_only=False)
    assert workbook.sheetnames == ["Summary", "SIG"]


def test_llcr_category_uses_maximum_group_sample_columns_and_disables_excess_cells(tmp_path) -> None:
    source = _projection()
    first = replace(source.sections[0], sample_count=2)
    second = replace(
        source.sections[0],
        confirmed_group_id="group-2",
        group_label="2",
        sample_count=1,
    )
    output_path = tmp_path / "mixed-sample-counts.xlsx"

    LlcrCrSpecializedRecordWorkbookGateway().write(
        output_path=output_path,
        projection=replace(source, sections=(first, second)),
    )

    workbook = load_workbook(output_path, data_only=False)
    sheet = workbook["SIG"]
    assert [sheet.cell(9, column).value for column in range(4, 6)] == ["1#", "2#"]
    assert [sheet.cell(9, column).value for column in range(11, 15)] == [
        "1#", "2#", "1#ΔR", "2#ΔR",
    ]
    assert sheet["A14"].value == "Group 2"
    assert sheet["K14"].value == '=IF(OR(D14="",$B$5=""),"",D14-$B$5)'
    assert sheet["L14"].value is None
    assert sheet["N14"].value is None
    assert sheet["E14"].fill.fgColor.rgb.endswith("E7E6E6")
    assert sheet["O14"].value == '=IF(COUNT(M14:M15)=0,"",MIN(M14:M15))'
    summary = workbook["Summary"]
    assert summary["A5"].value == "Group 2"
    assert summary["C5"].value == '=IF(\'SIG\'!O14="","",\'SIG\'!O14)'


def test_llcr_summary_places_requirement_parameters_over_category_statistics(tmp_path) -> None:
    source = _projection()
    signal = source.sections[0]
    power = replace(
        signal,
        category_id="ppc-2",
        category_label="Power",
        record_prefix="PWR",
        rows=(
            LlcrCrRecordRow(sample_index=1, contact_id="PWR1", contact_label="Power contact"),
            LlcrCrRecordRow(sample_index=1, contact_id="PWR2", contact_label="Power contact"),
        ),
    )
    output_path = tmp_path / "multiple-categories.xlsx"

    LlcrCrSpecializedRecordWorkbookGateway().write(
        output_path=output_path,
        projection=replace(
            source,
            sections=(signal, power),
            summary_parameter_labels=("Signal Contact", "Power Contact"),
        ),
    )

    workbook = load_workbook(output_path, data_only=False)
    assert workbook.sheetnames == ["Summary", "SIG", "PWR"]
    summary = workbook["Summary"]
    assert summary["C1"].value == "Signal Contact"
    assert summary["G1"].value == "Power Contact"
    assert summary["C3"].value == '=IF(\'SIG\'!L10="","",\'SIG\'!L10)'
    assert summary["G3"].value == '=IF(\'PWR\'!L10="","",\'PWR\'!L10)'


def test_llcr_summary_supports_single_measurement_groups_with_one_point(tmp_path) -> None:
    source = _projection()
    one_point_rows = (
        LlcrCrRecordRow(sample_index=1, contact_id="SIG1", contact_label="Signal contact"),
    )
    first = replace(
        source.sections[0],
        rows=one_point_rows,
        readings_per_sample=1,
        stages=(source.sections[0].stages[0],),
    )
    second = replace(
        first,
        confirmed_group_id="group-2",
        group_label="Group 2",
    )
    output_path = tmp_path / "single-measurement-groups.xlsx"

    LlcrCrSpecializedRecordWorkbookGateway().write(
        output_path=output_path,
        projection=replace(source, sections=(first, second)),
    )

    workbook = load_workbook(output_path, data_only=False)
    summary = workbook["Summary"]
    assert summary["B3"].value == "Initial LLCR"
    assert summary["B4"].value == "Initial LLCR"
    assert summary["C4"].value == '=IF(\'SIG\'!L11="","",\'SIG\'!L11)'
    assert summary["C4"].fill.fgColor.rgb.endswith("FFFFC8")


def test_llcr_without_delta_r_summarizes_corrected_measurements(tmp_path) -> None:
    source = _projection()
    output_path = tmp_path / "llcr-without-delta.xlsx"

    LlcrCrSpecializedRecordWorkbookGateway().write(
        output_path=output_path,
        projection=replace(source, delta_r_enabled=False),
    )

    workbook = load_workbook(output_path, data_only=False)
    sheet = workbook["SIG"]
    assert [sheet.cell(9, column).value for column in range(10, 15)] == [
        "1#", "Min", "Max", "Avg", "Stdev",
    ]
    assert all("ΔR" not in str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value)
    assert sheet["K10"].value == '=IF(COUNT(J10:J11)=0,"",MIN(J10:J11))'
    summary = workbook["Summary"]
    assert summary["B4"].value == "Final LLCR"
    assert summary["C3"].value == '=IF(\'SIG\'!K10="","",\'SIG\'!K10)'


def test_gateway_writes_cr_bulk_voltage_and_matrix_current_formula_without_delta(tmp_path) -> None:
    output_path = tmp_path / "cr-record.xlsx"
    source = _projection()
    cr_stages = tuple(
        replace(stage, test_item="CONTACT RESISTANCE (Power)", condition="10 A max", test_current_ampere="10")
        for stage in source.sections[0].stages
    )
    cr_section = replace(source.sections[0], record_type="cr", stages=cr_stages)
    projection = replace(source, record_type="cr", delta_r_enabled=False, sections=(cr_section,))

    LlcrCrSpecializedRecordWorkbookGateway().write(
        output_path=output_path, projection=projection,
    )

    workbook = load_workbook(output_path, data_only=False)
    assert workbook.sheetnames == ["Record Summary", "SIG"]
    sheet = workbook["SIG"]
    assert sheet["B8"].value == "Bulk Voltage (mV)"
    assert sheet["D16"].value == "Test current (A)"
    assert sheet["E16"].value == 10
    assert sheet["E18"].value == '=IF(OR(D18="",$E$16="",NOT(COUNTIFS($A$9:$A$10,$B18,$B$9:$B$10,"<>")>0)),"",(D18-VLOOKUP($B18,$A$9:$B$10,2,FALSE))/$E$16)'
    assert sheet["E24"].value == '=IF(COUNT(E18:E19)<2,"",STDEV(E18:E19))'
    assert all("ΔR" not in str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value)


def _projection() -> LlcrCrRecordProjection:
    section = LlcrCrRecordSection(
        record_type="llcr",
        confirmed_group_id="group-1",
        confirmed_row_id="row-1",
        step_sequence=2,
        step_suffix_note="",
        group_label="Group 1",
        source_step="2",
        sample_count=1,
        readings_per_sample=2,
        rows=(
            LlcrCrRecordRow(sample_index=1, contact_id="SIG1", contact_label="Signal contact"),
            LlcrCrRecordRow(sample_index=1, contact_id="SIG2", contact_label="Signal contact"),
        ),
        category_id="ppc-1",
        category_label="Signal",
        point_expression="1-2",
        stages=(
            LlcrCrRecordStage("Initial", "1", "row-1", "LLCR", "100 mA max", "", None),
            LlcrCrRecordStage("Final", "2", "row-2", "LLCR", "100 mA max", "", None),
        ),
        record_prefix="SIG",
    )
    return LlcrCrRecordProjection(
        project_id="project-1",
        confirmed_matrix_id="cmv-1",
        confirmed_revision=4,
        status="ready",
        sections=(section,),
        diagnostics=(),
        preview_fingerprint="fingerprint",
        record_type="llcr",
        point_profile_revision_id="profile-1",
        point_profile_revision_sequence=2,
        point_profile_fingerprint="profile-fingerprint",
        delta_r_enabled=True,
        ltr_number="DL-2025-11-073",
    )
