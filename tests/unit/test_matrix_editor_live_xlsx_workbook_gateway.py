from io import BytesIO

from openpyxl import load_workbook

from backend.application.matrix_editor_live_xlsx_export_service import (
    MatrixEditorLiveXlsxExportCell,
    MatrixEditorLiveXlsxExportGroup,
    MatrixEditorLiveXlsxExportProjection,
    MatrixEditorLiveXlsxExportRow,
    MatrixEditorLiveXlsxExportSchedule,
)
from backend.infrastructure.office.matrix_editor_live_xlsx_workbook_gateway import (
    MatrixEditorLiveXlsxWorkbookGateway,
)


def test_workbook_gateway_writes_reference_layout_and_true_blank_fee_cells():
    projection = MatrixEditorLiveXlsxExportProjection(
        groups=(MatrixEditorLiveXlsxExportGroup("g1", "G1", "Group 1", "5", "2.5 d"),),
        rows=(
            MatrixEditorLiveXlsxExportRow(
                "r1", "Thermal Shock", "4", "EIA", "-40/125", "Pass",
                (MatrixEditorLiveXlsxExportCell("g1", "1"),),
            ),
        ),
    )
    content = MatrixEditorLiveXlsxWorkbookGateway().render(projection)
    workbook = load_workbook(BytesIO(content), data_only=False)
    sheet = workbook["Sheet"]
    assert sheet.max_row == 5
    assert [cell.value for cell in sheet[1]] == [
        "Test Item", "Section", "Test Method", "Condition", "Requirement", "Group 1", "Notes"
    ]
    assert [sheet.cell(row, 1).value for row in range(2, 6)] == [
        "Thermal Shock", "Sample size", "Time", "Fee"
    ]
    assert sheet["F3"].value == "5"
    assert sheet["F4"].value == "2.5 d"
    assert all(sheet.cell(5, column).value is None for column in range(2, 8))
    assert sheet["A1"].fill.fgColor.rgb == "00CCCCCC"
    assert workbook.defined_names == {}


def test_workbook_gateway_embeds_versioned_fingerprint_metadata():
    projection = MatrixEditorLiveXlsxExportProjection(
        groups=(
            MatrixEditorLiveXlsxExportGroup(
                "g1", "G1", "Group 1", "5", "2.5 d", "Two reserves"
            ),
        ),
        rows=(
            MatrixEditorLiveXlsxExportRow(
                "r1", "Thermal Shock", "4", "EIA", "-40/125", "Pass",
                (MatrixEditorLiveXlsxExportCell("g1", "1"),), "0.5x",
            ),
        ),
        schedule=MatrixEditorLiveXlsxExportSchedule(post_test_buffer_days="1"),
    )

    content = MatrixEditorLiveXlsxWorkbookGateway().render(projection)
    workbook = load_workbook(BytesIO(content), data_only=False)
    metadata = workbook["__ConnLab_Metadata"]

    assert metadata.sheet_state == "veryHidden"
    assert metadata["A1"].value == "connlab.matrix.xlsx"
    assert metadata["B1"].value == "1"
    assert metadata["A2"].value == "visible_sha256"
    assert isinstance(metadata["B2"].value, str) and len(metadata["B2"].value) == 64
    assert metadata["A3"].value == "payload_json"
    assert '"day_expression":"0.5x"' in metadata["B3"].value
    assert metadata["A4"].value == "payload_sha256"
    assert isinstance(metadata["B4"].value, str) and len(metadata["B4"].value) == 64


def test_workbook_gateway_literalizes_formula_shaped_dynamic_text():
    formula = '=HYPERLINK("https://example.invalid","click")'
    projection = MatrixEditorLiveXlsxExportProjection(
        groups=(MatrixEditorLiveXlsxExportGroup("g1", "G1", formula, formula, "=2+2"),),
        rows=(
            MatrixEditorLiveXlsxExportRow(
                "r1", formula, "=1+1", "=SUM(1,2)", "=NOW()", "=A1",
                (MatrixEditorLiveXlsxExportCell("g1", formula),),
            ),
        ),
    )
    content = MatrixEditorLiveXlsxWorkbookGateway().render(projection)
    workbook = load_workbook(BytesIO(content), data_only=False)
    sheet = workbook["Sheet"]
    dynamic_cells = [
        sheet["F1"], sheet["A2"], sheet["B2"], sheet["C2"], sheet["D2"],
        sheet["E2"], sheet["F2"], sheet["F3"], sheet["F4"],
    ]
    assert [cell.value for cell in dynamic_cells] == [
        formula, formula, "=1+1", "=SUM(1,2)", "=NOW()", "=A1",
        formula, formula, "=2+2",
    ]
    assert all(cell.data_type != "f" for cell in dynamic_cells)
    assert all(cell.hyperlink is None for row in sheet.iter_rows() for cell in row)
    assert workbook.defined_names == {}
    assert workbook._external_links == []
    assert all(sheet.cell(5, column).value is None for column in range(2, 8))


def test_workbook_gateway_leaves_wrapped_rows_at_automatic_height():
    projection = MatrixEditorLiveXlsxExportProjection(
        groups=(MatrixEditorLiveXlsxExportGroup("g1", "G1", "Group 1", "5", "2.5 d"),),
        rows=(
            MatrixEditorLiveXlsxExportRow(
                "r1",
                "Cycling Temperature & Humidity",
                "8.2",
                "EIA-364-31",
                "Relative humidity and temperature: 25 ± 3 °C at 80 ± 5% relative humidity",
                "No damage and no discontinuity longer than one microsecond",
                (MatrixEditorLiveXlsxExportCell("g1", "7"),),
            ),
        ),
    )
    content = MatrixEditorLiveXlsxWorkbookGateway().render(projection)
    workbook = load_workbook(BytesIO(content), data_only=False)
    sheet = workbook["Sheet"]

    assert all(
        sheet.row_dimensions[row_number].height is None
        and not sheet.row_dimensions[row_number].customHeight
        for row_number in range(1, sheet.max_row + 1)
    )
    assert all(
        cell.alignment.wrap_text
        for row in sheet.iter_rows()
        for cell in row
    )
