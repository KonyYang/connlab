from __future__ import annotations

from openpyxl import load_workbook

from backend.application.test_status_workbook_projection import (
    TestStatusGroup,
    TestStatusRow,
    build_test_status_projection,
)
from backend.application.confirmed_matrix_test_status_workbook_generation_service import (
    ConfirmedMatrixTestStatusWorkbookGenerationService,
    GenerateConfirmedMatrixTestStatusWorkbookCommand,
)
from backend.infrastructure.office.test_status_workbook_gateway import (
    TestStatusWorkbookGateway,
)
from tests.unit.test_confirmed_matrix_test_record_preview_service import _snapshot


def test_workbook_preserves_matrix_group_values_and_vba_status_rows(tmp_path) -> None:
    projection = build_test_status_projection(
        groups=(
            TestStatusGroup("g1", "1", "5"),
            TestStatusGroup("g6", "6", "5+5(d)"),
        ),
        rows=(
            TestStatusRow("Examination", {"g1": "1,8", "g6": "1,7"}),
            TestStatusRow("LLCR", {"g1": "2,5,7", "g6": "3,5"}),
        ),
    )
    output_path = tmp_path / "DL-2026-08-001 test status.xlsx"

    TestStatusWorkbookGateway().write(
        output_path=output_path,
        projection=projection,
    )

    workbook = load_workbook(output_path, data_only=False)
    assert workbook.sheetnames == ["Test Status"]
    sheet = workbook["Test Status"]
    assert [sheet.cell(1, column).value for column in range(1, 4)] == ["TEST", "1", "6"]
    assert [sheet.cell(2, column).value for column in range(1, 4)] == [
        "Examination",
        "1,8",
        "1,7",
    ]
    assert [sheet.cell(4, column).value for column in range(1, 4)] == [
        "Sample size",
        "5",
        "5+5(d)",
    ]
    assert sheet["A5"].value == "Estimated completion date in Clarizen"
    assert sheet["B5"].value is None
    assert "B5:C5" in {str(item) for item in sheet.merged_cells.ranges}
    assert [sheet.cell(6, column).value for column in range(1, 4)] == [
        "Status",
        "No Start",
        "No Start",
    ]
    assert sheet["A1"].fill.fgColor.rgb == "FFC8C8C8"
    assert sheet["A4"].fill.fgColor.rgb == "FFADD8E6"
    assert sheet["C6"].fill.fgColor.rgb == "FFADD8E6"
    assert sheet.column_dimensions["A"].width == 30
    assert sheet.column_dimensions["B"].width == 15


def test_projection_rejects_duplicate_group_keys() -> None:
    try:
        build_test_status_projection(
            groups=(
                TestStatusGroup("g1", "1", "5"),
                TestStatusGroup("g1", "2", "3"),
            ),
            rows=(),
        )
    except ValueError as exc:
        assert str(exc) == "Test Status group keys must be unique."
    else:
        raise AssertionError("Expected duplicate Test Status group keys to be rejected.")


def test_confirmed_generation_uses_active_matrix_authority(tmp_path) -> None:
    snapshot = _snapshot()
    output_dir = tmp_path / "stage"
    service = ConfirmedMatrixTestStatusWorkbookGenerationService(
        confirmed_store=_ConfirmedStore(snapshot),
        writer=TestStatusWorkbookGateway(),
    )

    output = service.generate(
        GenerateConfirmedMatrixTestStatusWorkbookCommand(
            project_id="P1",
            output_dir=output_dir,
            target_name="DL-001 test status.xlsx",
        )
    )

    workbook = load_workbook(output, data_only=False)
    sheet = workbook["Test Status"]
    assert [sheet.cell(1, column).value for column in range(1, 4)] == ["TEST", "1", "2"]
    assert [sheet.cell(2, column).value for column in range(1, 4)] == [
        "Visual",
        "1,2(a)",
        "3",
    ]
    assert [sheet.cell(4, column).value for column in range(1, 4)] == [
        "Sample size",
        "5",
        "6",
    ]


class _ConfirmedStore:
    def __init__(self, snapshot) -> None:
        self.snapshot = snapshot

    def get_active_by_project(self, project_id: str):
        return self.snapshot if self.snapshot.version.project_id == project_id else None
