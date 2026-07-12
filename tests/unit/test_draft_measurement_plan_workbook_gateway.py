from pathlib import Path

from openpyxl import load_workbook

from backend.application.draft_measurement_plan_workbook_projection import (
    build_draft_measurement_plan_workbook_projection,
)
from backend.infrastructure.office.draft_measurement_plan_workbook_gateway import (
    DraftMeasurementPlanWorkbookGateway,
)


def test_gateway_writes_macro_free_needs_review_banner_and_metadata(tmp_path: Path) -> None:
    workspace = _workspace()
    workspace["impacts"] = [{"severity": "review_required", "resolution_state": "open"}]
    projection = build_draft_measurement_plan_workbook_projection(workspace)
    path = tmp_path / "draft.xlsx"

    DraftMeasurementPlanWorkbookGateway().write(output_path=path, projection=projection)

    workbook = load_workbook(path, keep_vba=False)
    assert workbook.sheetnames == ["Record Summary", "LLCR Record", "CR Record"]
    assert workbook["Record Summary"]["A1"].value == "NEEDS REVIEW"
    assert workbook["Record Summary"]["B4"].value == "draft-1"
    assert workbook["Record Summary"]["B5"].value == "f"
    assert workbook["Record Summary"]["B6"].value == "matrix-1"
    assert workbook["Record Summary"]["B7"].value == "matrix"
    assert workbook["Record Summary"]["A10"].value == "Layout version"
    assert workbook["Record Summary"]["B10"].value == "LLCR_CR_RECORD_LAYOUT_V1"
    assert workbook["Record Summary"]["A15"].value.startswith("DRAFT ONLY")
    assert workbook["Record Summary"]["A18"].value == "LLCR"
    assert workbook["LLCR Record"]["A1"].value.startswith("NEEDS REVIEW")
    assert workbook["CR Record"]["A1"].value.startswith("NEEDS REVIEW")
    assert workbook["LLCR Record"]["G7"].value.startswith("=IF(COUNT(")
    assert workbook["LLCR Record"].column_dimensions["K"].width == 28
    assert not workbook.vba_archive


def _workspace() -> dict[str, object]:
    return {
        "project_id": "P-1", "editable_revision_id": "draft-1", "editable_revision_state": "draft", "editable_revision_fingerprint": "f", "revision": {"revision_id": "draft-1", "revision_sequence": 1},
        "matrix_binding": {"base_confirmed_matrix_id": "matrix-1", "base_matrix_revision": 1, "matrix_binding_fingerprint": "matrix"},
        "targets": [{"contact_kind": "llcr", "eligible": True, "included": True, "group_label": "Group", "step_sequence": 1, "step_suffix_note": "", "sample_quantity_expression": "1", "readings_per_sample": 1, "families": [{"included": True, "count_per_sample": 1, "record_label": "High Power", "record_prefix": "HP"}]}],
        "impacts": [],
    }
