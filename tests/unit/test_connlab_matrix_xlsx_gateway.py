from pathlib import Path
import json

from openpyxl import Workbook, load_workbook

from backend.application.matrix_editor_live_xlsx_export_service import (
    MatrixEditorLiveXlsxExportCell,
    MatrixEditorLiveXlsxExportGroup,
    MatrixEditorLiveXlsxExportProjection,
    MatrixEditorLiveXlsxExportRow,
    MatrixEditorLiveXlsxExportSchedule,
)
from backend.infrastructure.office.connlab_matrix_xlsx_gateway import (
    ConnLabMatrixXlsxGateway,
)
from backend.infrastructure.office.matrix_editor_live_xlsx_workbook_gateway import (
    MatrixEditorLiveXlsxWorkbookGateway,
)


def _write_visible_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet"
    sheet.append(
        [
            "Test Item",
            "Section",
            "Test Method",
            "Condition",
            "Requirement",
            "Group 1",
            "Group 2",
            "Notes",
        ]
    )
    sheet.append(["Visual", "5.1", "EIA-364-18", "10x", "No damage", "1", "", None])
    sheet.append(["Thermal Shock", "8.1", "EIA-364-32", "-40/125", "Pass", "2", "1a", None])
    sheet.append(["Sample size", None, None, None, None, "5", "5+5(d)", None])
    sheet.append(["Time", None, None, None, None, "2 d", "1 d", None])
    sheet.append(["Fee", None, None, None, None, None, None, None])
    workbook.save(path)


def _projection() -> MatrixEditorLiveXlsxExportProjection:
    return MatrixEditorLiveXlsxExportProjection(
        groups=(
            MatrixEditorLiveXlsxExportGroup(
                "g1", "group_primary", "Group 1", "5", "2 d", "Extra 2 samples"
            ),
        ),
        rows=(
            MatrixEditorLiveXlsxExportRow(
                "r1",
                "Thermal Shock",
                "8.1",
                "EIA-364-32",
                "-40/125",
                "Pass",
                (MatrixEditorLiveXlsxExportCell("g1", "1"),),
                "2.5x",
            ),
        ),
        schedule=MatrixEditorLiveXlsxExportSchedule(
            post_test_buffer_days="2",
            sample_received_date="2026-08-01",
            planned_test_start_date="2026-08-02",
            planned_test_complete_date="2026-08-05",
            estimated_completion_date="2026-08-07",
        ),
    )


def test_reads_legacy_connlab_visible_table_and_defaults_day_to_zero(tmp_path: Path) -> None:
    path = tmp_path / "legacy.xlsx"
    _write_visible_workbook(path)

    result = ConnLabMatrixXlsxGateway().read(path)

    assert result.blockers == ()
    assert [group.group_label for group in result.groups] == ["Group 1", "Group 2"]
    assert result.groups[1].sample_quantity_expression == "5+5(d)"
    assert [row.day_expression for row in result.rows] == ["0", "0"]
    assert result.rows[1].group_tokens == {"group_1": "2", "group_2": "1a"}
    assert any("default" in warning.lower() and "Day" in warning for warning in result.warnings)


def test_visible_group_keys_preserve_numeric_label_gaps(tmp_path: Path) -> None:
    path = tmp_path / "group-gap.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "Test Item",
            "Section",
            "Test Method",
            "Condition",
            "Requirement",
            "5",
            "7",
            "Notes",
        ]
    )
    sheet.append(["Visual", "5.1", "EIA-364-18", "10x", "No damage", "1", "1,5", None])
    sheet.append(["Sample size", None, None, None, None, "5", "3", None])
    sheet.append(["Time", None, None, None, None, "0 d", "0 d", None])
    sheet.append(["Fee", None, None, None, None, None, None, None])
    workbook.save(path)

    result = ConnLabMatrixXlsxGateway().read(path)

    assert result.blockers == ()
    assert [(group.group_key, group.group_label) for group in result.groups] == [
        ("group_5", "5"),
        ("group_7", "7"),
    ]
    assert result.rows[0].group_tokens == {"group_5": "1", "group_7": "1,5"}


def test_round_trip_prefers_matching_hidden_metadata(tmp_path: Path) -> None:
    content = MatrixEditorLiveXlsxWorkbookGateway().render(_projection())
    path = tmp_path / "round-trip.xlsx"
    path.write_bytes(content)

    workbook = load_workbook(path, data_only=False)
    assert workbook["__ConnLab_Metadata"].sheet_state == "veryHidden"

    result = ConnLabMatrixXlsxGateway().read(path)

    assert result.blockers == ()
    assert result.warnings == ()
    assert result.groups[0].group_key == "group_primary"
    assert result.groups[0].sample_note == "Extra 2 samples"
    assert result.rows[0].day_expression == "2.5x"
    assert result.schedule == {
        "post_test_buffer_days": "2",
        "sample_received_date": "2026-08-01",
        "planned_test_start_date": "2026-08-02",
        "planned_test_complete_date": "2026-08-05",
        "estimated_completion_date": "2026-08-07",
    }


def test_stale_hidden_metadata_falls_back_to_visible_values(tmp_path: Path) -> None:
    path = tmp_path / "edited.xlsx"
    path.write_bytes(MatrixEditorLiveXlsxWorkbookGateway().render(_projection()))
    workbook = load_workbook(path, data_only=False)
    workbook["Sheet"]["A2"] = "Edited Thermal Shock"
    workbook.save(path)

    result = ConnLabMatrixXlsxGateway().read(path)

    assert result.blockers == ()
    assert result.rows[0].test_item == "Edited Thermal Shock"
    assert result.rows[0].day_expression == "0"
    assert any("details" in warning.lower() and "visible" in warning.lower() for warning in result.warnings)


def test_modified_hidden_payload_is_not_trusted(tmp_path: Path) -> None:
    path = tmp_path / "metadata-edited.xlsx"
    path.write_bytes(MatrixEditorLiveXlsxWorkbookGateway().render(_projection()))
    workbook = load_workbook(path, data_only=False)
    metadata = workbook["__ConnLab_Metadata"]
    payload = json.loads(metadata["B3"].value)
    payload["rows"][0]["day_expression"] = "999"
    metadata["B3"] = json.dumps(payload)
    workbook.save(path)

    result = ConnLabMatrixXlsxGateway().read(path)

    assert result.rows[0].day_expression == "0"
    assert any("details" in warning.lower() for warning in result.warnings)


def test_rejects_non_connlab_xlsx_structure(tmp_path: Path) -> None:
    path = tmp_path / "other.xlsx"
    workbook = Workbook()
    workbook.active.append(["Name", "Value"])
    workbook.save(path)

    result = ConnLabMatrixXlsxGateway().read(path)

    assert result.groups == ()
    assert any("ConnLab" in blocker for blocker in result.blockers)
