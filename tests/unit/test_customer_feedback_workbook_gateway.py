from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from backend.infrastructure.office.customer_feedback_workbook_gateway import (
    CustomerFeedbackWorkbookGateway,
    CustomerFeedbackWorkbookGatewayError,
)


def test_customer_feedback_workbook_gateway_copies_template_without_overwriting_source(
    tmp_path: Path,
) -> None:
    template = tmp_path / "E-4243_D Customer Feedback Form.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "LTR Number"
    sheet["A2"] = "Product Description"
    sheet["A3"] = "Requestor"
    workbook.save(template)
    output = tmp_path / "generated" / "feedback.xlsx"

    result_path, warnings = CustomerFeedbackWorkbookGateway().generate(
        template_path=template,
        output_path=output,
        identity={
            "ltr_number": "DL-2026-05-003",
            "product_name": "Connector",
            "requestor": "MP Cao",
        },
    )

    assert result_path == output
    generated = load_workbook(output)
    sheet = generated.active
    assert sheet["B1"].value == "DL-2026-05-003"
    assert sheet["B2"].value == "Connector"
    assert sheet["B3"].value == "MP Cao"
    assert warnings == ()
    source = load_workbook(template)
    assert source.active["B1"].value is None


def test_customer_feedback_workbook_gateway_fills_sample_compatible_header_offsets(
    tmp_path: Path,
) -> None:
    template = tmp_path / "E-4243_D Customer Feedback Form.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Customer Feedback Form"
    sheet["A7"] = "Customer Name"
    sheet["D7"] = "Telephone No."
    sheet["F7"] = "Site"
    sheet["A9"] = "Project Details\n( if applicable)"
    sheet["F9"] = "Work Request No."
    sheet["A11"] = "From Date\n(mm/dd/yy)"
    sheet["D11"] = "To Date\n(mm/dd/yy)"
    sheet["A13"] = "GES Team"
    workbook.save(template)
    output = tmp_path / "generated" / "feedback.xlsx"

    result_path, warnings = CustomerFeedbackWorkbookGateway().generate(
        template_path=template,
        output_path=output,
        identity={
            "ltr_number": "DL-BI",
            "product_name": "Connector BI Qualification",
            "requestor": "Requester BI",
            "phone": "12345",
            "location": "Dongguan",
            "received_date": "20 Jun 2026",
            "estimated_completion_date": "02 Jul 2026",
            "lab": "Dongguan Lab",
        },
    )

    assert result_path == output
    generated = load_workbook(output)
    sheet = generated["Customer Feedback Form"]
    assert sheet["C7"].value == "Requester BI"
    assert sheet["E7"].value == "12345"
    assert sheet["I7"].value == "Dongguan"
    assert sheet["C9"].value == "Connector BI Qualification"
    assert sheet["I9"].value == "DL-BI"
    assert sheet["C11"].value == "20 Jun 2026"
    assert sheet["E11"].value == "02 Jul 2026"
    assert sheet["C13"].value == "Dongguan Lab"
    assert warnings == ()


def test_customer_feedback_workbook_gateway_blocks_missing_required_identity_anchor(
    tmp_path: Path,
) -> None:
    template = tmp_path / "E-4243_D Customer Feedback Form.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "Customer Name"
    workbook.save(template)

    with pytest.raises(CustomerFeedbackWorkbookGatewayError, match="Work Request No"):
        CustomerFeedbackWorkbookGateway().generate(
            template_path=template,
            output_path=tmp_path / "generated" / "feedback.xlsx",
            identity={
                "ltr_number": "DL-BI",
                "product_name": "Connector BI Qualification",
            },
        )


def test_customer_feedback_workbook_gateway_rejects_non_xlsx_template(tmp_path: Path) -> None:
    template = tmp_path / "E-4243.xls"
    template.write_bytes(b"template")

    with pytest.raises(CustomerFeedbackWorkbookGatewayError, match=".xlsx"):
        CustomerFeedbackWorkbookGateway().generate(
            template_path=template,
            output_path=tmp_path / "out.xlsx",
            identity={},
        )


def test_customer_feedback_workbook_gateway_rejects_output_equal_to_template(
    tmp_path: Path,
) -> None:
    template = tmp_path / "E-4243.xlsx"
    template.write_bytes(b"template")

    with pytest.raises(CustomerFeedbackWorkbookGatewayError, match="must not overwrite"):
        CustomerFeedbackWorkbookGateway().generate(
            template_path=template,
            output_path=template,
            identity={},
        )
