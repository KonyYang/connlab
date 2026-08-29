from __future__ import annotations

from dataclasses import replace
from decimal import Decimal

from openpyxl import load_workbook

from backend.application.confirmed_matrix_llcr_cr_record_projection import (
    LlcrCrRecordProjection,
    LlcrCrRecordDiagnostic,
    LlcrCrRecordRow,
    LlcrCrRecordSection,
    LlcrCrRecordStage,
)
from backend.infrastructure.office.llcr_cr_specialized_record_workbook_gateway import (
    LlcrCrSpecializedRecordWorkbookGateway,
)
from backend.infrastructure.office.llcr_result_workbook_gateway import (
    LlcrResultWorkbookGateway,
)


def test_inspect_reads_original_precision_and_maps_stages_to_matrix_steps(tmp_path) -> None:
    source = tmp_path / "completed-llcr.xlsx"
    projection = _projection()
    LlcrCrSpecializedRecordWorkbookGateway().write(
        output_path=source,
        projection=projection,
    )
    workbook = load_workbook(source, data_only=False)
    sheet = workbook["SIG"]
    sheet["B2"], sheet["B3"], sheet["B4"] = Decimal("0.05"), Decimal("0.05"), Decimal("0.05")
    sheet["D10"], sheet["D11"] = Decimal("0.219"), Decimal("0.248")
    sheet["D12"], sheet["D13"] = Decimal("0.220"), Decimal("0.270")
    workbook.save(source)

    inspection = LlcrResultWorkbookGateway().inspect(
        source_path=source,
        projection=projection,
    )

    assert inspection.parser_profile_version == "connlab-llcr-macro-v1"
    assert inspection.detected_sheets == ("Summary", "SIG")
    assert inspection.diagnostics == ()
    initial, final = inspection.entries
    assert (initial.group_label, initial.matrix_step_sequence, initial.stage) == (
        "1",
        2,
        "initial",
    )
    assert [value.value for value in initial.measurements] == [
        Decimal("0.169"),
        Decimal("0.198"),
    ]
    assert [value.raw_value for value in initial.measurements] == [
        Decimal("0.219"),
        Decimal("0.248"),
    ]
    assert [value.raw_source_cell for value in initial.measurements] == ["D10", "D11"]
    assert initial.summary_max == Decimal("0.198")
    assert initial.provisional_outcome == "pass"
    assert final.matrix_step_sequence == 4
    assert [value.value for value in final.measurements] == [
        Decimal("0.001"),
        Decimal("0.022"),
    ]
    assert final.summary_max == Decimal("0.022")
    assert final.provisional_outcome == "pass"
    assert final.source_range == "SIG!K12:K13"


def test_inspect_returns_blocking_diagnostic_for_a_corrupt_xlsx(tmp_path) -> None:
    source = tmp_path / "corrupt-llcr.xlsx"
    source.write_bytes(b"not-an-openxml-workbook")

    inspection = LlcrResultWorkbookGateway().inspect(
        source_path=source,
        projection=_projection(),
    )

    assert inspection.entries == ()
    assert any(
        item.code == "unsupported_workbook_structure"
        and item.severity == "blocked"
        for item in inspection.diagnostics
    )


def test_inspect_blocks_duplicate_matrix_mapping_before_reading_results(tmp_path) -> None:
    source = tmp_path / "duplicate-mapping-llcr.xlsx"
    projection = _projection()
    LlcrCrSpecializedRecordWorkbookGateway().write(
        output_path=source,
        projection=projection,
    )

    inspection = LlcrResultWorkbookGateway().inspect(
        source_path=source,
        projection=replace(
            projection,
            sections=(projection.sections[0], projection.sections[0]),
        ),
    )

    assert inspection.entries == ()
    assert any(item.code == "duplicate_mapping" for item in inspection.diagnostics)


def test_inspect_blocks_confirmation_when_required_measurement_is_missing(tmp_path) -> None:
    source = tmp_path / "incomplete-llcr.xlsx"
    projection = _projection()
    LlcrCrSpecializedRecordWorkbookGateway().write(
        output_path=source,
        projection=projection,
    )
    workbook = load_workbook(source, data_only=False)
    sheet = workbook["SIG"]
    sheet["B2"], sheet["B3"], sheet["B4"] = 0.05, 0.05, 0.05
    sheet["D10"], sheet["D11"] = 0.219, 0.248
    sheet["D12"] = 0.220
    workbook.save(source)

    inspection = LlcrResultWorkbookGateway().inspect(
        source_path=source,
        projection=projection,
    )

    assert inspection.entries == ()
    assert any(item.code == "required_measurement_missing" for item in inspection.diagnostics)
    assert any(item.severity == "blocked" for item in inspection.diagnostics)


def test_inspect_selects_initial_and_delta_clauses_from_combined_requirement(tmp_path) -> None:
    source = tmp_path / "combined-requirement-llcr.xlsx"
    projection = _projection(
        initial_requirement="Initial ≤ 0.25 mΩ; ΔR ≤ 0.17 mΩ",
        final_requirement="Initial ≤ 0.25 mΩ; ΔR ≤ 0.17 mΩ",
    )
    LlcrCrSpecializedRecordWorkbookGateway().write(
        output_path=source,
        projection=projection,
    )
    workbook = load_workbook(source, data_only=False)
    sheet = workbook["SIG"]
    sheet["B2"], sheet["B3"], sheet["B4"] = 0.05, 0.05, 0.05
    sheet["D10"], sheet["D11"] = 0.219, 0.248
    sheet["D12"], sheet["D13"] = 0.220, 0.270
    workbook.save(source)

    inspection = LlcrResultWorkbookGateway().inspect(
        source_path=source,
        projection=projection,
    )

    assert inspection.diagnostics == ()
    initial, final = inspection.entries
    assert initial.requirement == "Initial ≤ 0.25 mΩ"
    assert initial.requirement_limit == Decimal("0.25")
    assert final.requirement == "ΔR ≤ 0.17 mΩ"
    assert final.requirement_limit == Decimal("0.17")


def test_inspect_blocks_delta_stage_without_delta_requirement(tmp_path) -> None:
    source = tmp_path / "missing-delta-requirement-llcr.xlsx"
    projection = _projection(
        initial_requirement="Initial ≤0.25mΩ",
        final_requirement="Initial ≤0.25mΩ",
    )
    LlcrCrSpecializedRecordWorkbookGateway().write(
        output_path=source,
        projection=projection,
    )
    workbook = load_workbook(source, data_only=False)
    sheet = workbook["SIG"]
    sheet["B2"], sheet["B3"], sheet["B4"] = 0.05, 0.05, 0.05
    sheet["D10"], sheet["D11"] = 0.219, 0.248
    sheet["D12"], sheet["D13"] = 0.220, 0.270
    workbook.save(source)

    inspection = LlcrResultWorkbookGateway().inspect(
        source_path=source,
        projection=projection,
    )

    assert inspection.entries == ()
    assert any(
        item.code == "requirement_not_interpretable"
        and item.group_label == "1"
        and item.step_token == "4"
        for item in inspection.diagnostics
    )


def test_inspect_supports_strict_less_than_and_converts_micro_ohms(tmp_path) -> None:
    source = tmp_path / "strict-comparator-llcr.xlsx"
    projection = _projection(
        initial_requirement="Initial < 250 µΩ",
        final_requirement="ΔR < 170 µΩ",
    )
    LlcrCrSpecializedRecordWorkbookGateway().write(
        output_path=source,
        projection=projection,
    )
    workbook = load_workbook(source, data_only=False)
    sheet = workbook["SIG"]
    sheet["B2"], sheet["B3"], sheet["B4"] = 0.05, 0.05, 0.05
    sheet["D10"], sheet["D11"] = 0.219, 0.248
    sheet["D12"], sheet["D13"] = 0.220, 0.270
    workbook.save(source)

    inspection = LlcrResultWorkbookGateway().inspect(
        source_path=source,
        projection=projection,
    )

    assert inspection.diagnostics == ()
    initial, final = inspection.entries
    assert initial.requirement_comparator == "<"
    assert initial.requirement_limit == Decimal("0.250")
    assert initial.provisional_outcome == "pass"
    assert final.requirement_comparator == "<"
    assert final.requirement_limit == Decimal("0.170")
    assert final.provisional_outcome == "pass"


def test_inspect_surfaces_partial_projection_omissions_as_blockers(tmp_path) -> None:
    source = tmp_path / "partial-compatible-llcr.xlsx"
    projection = replace(
        _projection(),
        status="partial_compatible",
        diagnostics=(
            LlcrCrRecordDiagnostic(
                code="measurement_plan_omission",
                severity="review_required",
                message="One LLCR target is omitted from the effective measurement plan.",
            ),
        ),
        omission_diagnostics=("One LLCR target is omitted.",),
    )
    LlcrCrSpecializedRecordWorkbookGateway().write(
        output_path=source,
        projection=projection,
    )
    workbook = load_workbook(source, data_only=False)
    sheet = workbook["SIG"]
    sheet["B2"], sheet["B3"], sheet["B4"] = 0.05, 0.05, 0.05
    sheet["D10"], sheet["D11"] = 0.219, 0.248
    sheet["D12"], sheet["D13"] = 0.220, 0.270
    workbook.save(source)

    inspection = LlcrResultWorkbookGateway().inspect(
        source_path=source,
        projection=projection,
    )

    assert len(inspection.entries) == 2
    assert any(
        item.code == "measurement_plan_omission"
        and item.severity == "blocked"
        for item in inspection.diagnostics
    )


def _projection(
    *,
    initial_requirement: str = "Initial ≤0.25mΩ",
    final_requirement: str = "ΔR ≤0.17mΩ",
) -> LlcrCrRecordProjection:
    stages = (
        LlcrCrRecordStage(
            label="Initial",
            source_step="2",
            confirmed_row_id="row-llcr",
            test_item="LLCR",
            condition="100 mA max",
            requirement=initial_requirement,
            test_current_ampere="0.1",
        ),
        LlcrCrRecordStage(
            label="Final",
            source_step="4",
            confirmed_row_id="row-llcr",
            test_item="LLCR",
            condition="100 mA max",
            requirement=final_requirement,
            test_current_ampere="0.1",
        ),
    )
    rows = (
        LlcrCrRecordRow(1, "SIG1", "Signal contact"),
        LlcrCrRecordRow(1, "SIG2", "Signal contact"),
    )
    section = LlcrCrRecordSection(
        record_type="llcr",
        confirmed_group_id="group-1",
        confirmed_row_id="row-llcr",
        step_sequence=2,
        step_suffix_note="",
        group_label="1",
        source_step="2, 4",
        sample_count=1,
        readings_per_sample=2,
        rows=rows,
        category_id="category-sig",
        category_label="Signal contact",
        point_expression="SIG1-SIG2",
        stages=stages,
        record_prefix="SIG",
    )
    return LlcrCrRecordProjection(
        project_id="P1",
        confirmed_matrix_id="matrix-1",
        confirmed_revision=3,
        status="ready",
        sections=(section,),
        diagnostics=(),
        preview_fingerprint="projection-hash",
        ltr_number="DL-001",
        record_type="llcr",
        delta_r_enabled=True,
    )
